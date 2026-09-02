import calendar as py_calendar
import copy
import html
import json
import logging
import re
import urllib.error
import urllib.request
import uuid
from datetime import date as date_cls
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.db.models import Count, Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.dateparse import parse_time

from src import metrics, payroll
from src import parser as attendance_parser

from .forms import UploadForm
from .importer import import_dataframe, import_file
from .models import (
    AttendanceRecord, CashRegisterEntry, CashWithdrawal, Department, EarlyClosureDay, Employee, EmploymentPeriod,
    LeaveLedgerEntry, MonthLock, SalaryAdjustment, SpecialDay, UploadBatch,
)

logger = logging.getLogger(__name__)


def _error(request, text: str) -> None:
    """messages.error() plus a matching log line — a user-facing error
    should always leave a trace in the log, not just a one-time flash
    message that's gone once the page reloads. Use this instead of
    messages.error() directly unless the call site already logs its own
    more detailed line (e.g. logger.exception with a traceback)."""
    messages.error(request, text)
    logger.warning("%s [%s, user=%s]", text, request.path, request.user)

# Mon..Sun abbreviations for the grid's day-of-week header — Thursday and
# Sunday get two letters (TH/SU) instead of just T/S, so they aren't
# ambiguous with Tuesday and Saturday.
_DOW_LABELS = {0: "M", 1: "T", 2: "W", 3: "TH", 4: "F", 5: "S", 6: "SU"}


def _parse_month_date(date_param: str | None, default: date_cls) -> date_cls:
    """Parses the "date" query param used to pick a target month across
    the dashboard and OT Details report — falls back to `default` for a
    missing or malformed value instead of letting a bad bookmarked/shared
    link 500."""
    if not date_param:
        return default
    try:
        return date_cls.fromisoformat(date_param)
    except ValueError:
        return default


@login_required
def home_view(request):
    """Landing page — the app's root URL. Just a branded splash with links
    into the three real pages (Upload, Attendance, Holiday Calendar)."""
    return render(request, "attendance/home.html")


# Uploaded files are staged here (as <32-hex-token>.<ext>) between the
# initial "parse" step and the "confirm" step of upload_view, so the
# employee checklist can be built and reviewed before anything touches the
# database. Never committed — see .gitignore.
_UPLOAD_STAGING_DIR = Path(settings.BASE_DIR) / "upload_staging"
_STAGED_UPLOAD_NAME_RE = re.compile(r"^[0-9a-f]{32}\.(xlsx|xlsm|xls|csv)$")


def _staged_upload_path(staged_name: str) -> Path | None:
    """Resolve a staged-upload filename from a form field back to its path
    on disk, rejecting anything that isn't exactly the token format this
    view generates (blocks path traversal via a hand-crafted field)."""
    if not staged_name or not _STAGED_UPLOAD_NAME_RE.match(staged_name):
        return None
    return _UPLOAD_STAGING_DIR / staged_name


def _cleanup_staged_upload(staged_name: str) -> None:
    path = _staged_upload_path(staged_name)
    if path and path.exists():
        path.unlink(missing_ok=True)


@login_required
def upload_view(request):
    if request.method == "POST":
        action = request.POST.get("action", "parse")

        if action == "cancel":
            _cleanup_staged_upload(request.POST.get("staged_name", ""))
            return redirect("upload")

        if action == "confirm":
            staged_name = request.POST.get("staged_name", "")
            original_name = request.POST.get("file_name") or "upload"
            selected_codes = set(request.POST.getlist("emp_codes"))
            staged_path = _staged_upload_path(staged_name)
            if not staged_path or not staged_path.exists():
                messages.error(request, "That upload has expired — please upload the file again.")
                return redirect("upload")
            try:
                raw = attendance_parser.load_file(str(staged_path))
                daily = attendance_parser.normalize(raw)
                daily = daily[daily["emp_code"].astype(str).isin(selected_codes)].reset_index(drop=True)
                batch = import_dataframe(daily, file_name=original_name)
                messages.success(
                    request,
                    f"Imported {batch.row_count} rows for {len(selected_codes)} selected "
                    f"employee(s), {batch.period_start} to {batch.period_end}.",
                )
                logger.info(
                    "Upload imported (filtered): %s rows, %s employees selected, file=%s",
                    batch.row_count, len(selected_codes), original_name,
                )
            except Exception as exc:  # noqa: BLE001 - surface any parse/import error to HR
                messages.error(request, f"Import failed: {exc}")
                logger.exception("Filtered upload import failed for file=%s", original_name)
            finally:
                _cleanup_staged_upload(staged_name)
            return redirect("upload")

        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded = request.FILES["file"]
            suffix = Path(uploaded.name).suffix.lower().lstrip(".")
            if suffix not in {"xlsx", "xlsm", "xls", "csv"}:
                messages.error(request, "Unsupported file type — expected .xlsx, .xlsm, .xls, or .csv.")
                return redirect("upload")

            _UPLOAD_STAGING_DIR.mkdir(exist_ok=True)
            staged_name = f"{uuid.uuid4().hex}.{suffix}"
            staged_path = _UPLOAD_STAGING_DIR / staged_name
            with open(staged_path, "wb") as fh:
                for chunk in uploaded.chunks():
                    fh.write(chunk)

            try:
                raw = attendance_parser.load_file(str(staged_path))
                daily = attendance_parser.normalize(raw)
            except Exception as exc:  # noqa: BLE001 - surface any parse error to HR
                messages.error(request, f"Could not read that file: {exc}")
                logger.exception("Upload parse failed for file=%s", uploaded.name)
                staged_path.unlink(missing_ok=True)
                return redirect("upload")

            employees = (
                daily[["emp_code", "emp_name"]]
                .drop_duplicates()
                .sort_values("emp_name")
                .to_dict("records")
            )
            recent_batches = UploadBatch.objects.all()[:10]
            return render(request, "attendance/upload.html", {
                "form": UploadForm(),
                "batches": recent_batches,
                "pending": {
                    "staged_name": staged_name,
                    "file_name": uploaded.name,
                    "employees": employees,
                    "row_count": len(daily),
                },
            })
    else:
        form = UploadForm()

    recent_batches = UploadBatch.objects.all()[:10]
    return render(request, "attendance/upload.html", {"form": form, "batches": recent_batches})


def _special_days_and_downgrades() -> tuple[dict, dict, dict]:
    """(special_days, downgraded, skip) for metrics.apply_special_days —
    one query for every SpecialDay plus its downgraded_employees, shared
    by every caller (Dashboard, OT Details, Leave Ledger) so they can't
    drift on which employees a given holiday's downgrade to plain
    Holiday applies to.

    skip additionally excludes every Contractor-department employee
    entirely from Paid Holiday dates — a blanket policy (contractors
    aren't entitled to paid holidays at all), so their day computes
    completely normally there instead of getting any Holiday/Paid
    Holiday treatment, same as an ordinary working day."""
    special_days: dict = {}
    downgraded: dict = {}
    for sd in SpecialDay.objects.prefetch_related("downgraded_employees"):
        special_days[sd.date] = sd.day_type
        codes = {emp.code for emp in sd.downgraded_employees.all()}
        if codes:
            downgraded[sd.date] = codes

    contractor_codes = set(
        Employee.objects.filter(department__name__iexact="Contractor").values_list("code", flat=True)
    )
    skip: dict = {}
    if contractor_codes:
        for d, day_type in special_days.items():
            if day_type == SpecialDay.PAID_HOLIDAY:
                skip[d] = contractor_codes

    return special_days, downgraded, skip


def _early_closure_hours() -> dict:
    """date -> expected full-day hours (from EarlyClosureDay), for
    metrics.is_short_hours/permission_hours_by_employee/
    month_attendance_view — a date with no entry here keeps their
    default 8.5h."""
    return {
        ec.date: float(ec.full_day_hours)
        for ec in EarlyClosureDay.objects.all()
    }


def _load_daily_data() -> pd.DataFrame:
    """Loads all attendance records into a DataFrame shaped for src/metrics.py.

    Excludes any row whose date falls outside every one of that
    employee's EmploymentPeriod ranges (employees with no periods
    recorded at all are left alone — same leniency EmployeeQuerySet's
    active_on()/active_during() already use). This matters because the
    eSSL device export still lists a terminated employee's ID (it
    doesn't know they left) and auto-marks them Absent every day, so a
    re-upload can silently create a real AttendanceRecord for someone
    long gone — active_on()/active_during() correctly keep such people
    out of the Attendance/OT/Salary employee *lists*, but without this
    filter here their leftover Absent row would still show up in every
    page built from this data (Dashboard, OT Details, Salary...), since
    none of them re-derive "was this person actually employed on this
    date" themselves."""
    rows = AttendanceRecord.objects.select_related("employee", "employee__department").values(
        "employee__code", "employee__name", "employee__department__name", "employee__designation",
        "employee__category", "employee__subcategory", "employee__company",
        "date", "shift", "time_in", "time_out", "work_hours", "ot_hours", "status", "manually_edited",
        "manually_edited_fields",
    )
    df = pd.DataFrame.from_records(rows)
    if df.empty:
        return df
    df = df.rename(columns={
        "employee__code": "emp_code", "employee__name": "emp_name",
        "employee__department__name": "department", "employee__designation": "designation",
        "employee__category": "category", "employee__subcategory": "subcategory",
        "employee__company": "company",
    })
    df["department"] = df["department"].fillna("Unassigned")
    df["date"] = pd.to_datetime(df["date"])

    periods_by_code: dict[str, list[tuple]] = {}
    for p in EmploymentPeriod.objects.values("employee__code", "start_date", "end_date"):
        periods_by_code.setdefault(p["employee__code"], []).append((p["start_date"], p["end_date"]))

    def _covered(emp_code: str, dt) -> bool:
        periods = periods_by_code.get(emp_code)
        if not periods:
            return True
        d = dt.date()
        return any(start <= d and (end is None or d <= end) for start, end in periods)

    covered_mask = [
        _covered(row.emp_code, row.date) for row in df.itertuples(index=False)
    ]
    return df[covered_mask].reset_index(drop=True)


def _ot_payable_table(shift_ot_table: pd.DataFrame, staff_codes: set) -> pd.DataFrame:
    """shift_ot_table (metrics.overtime_view's output) with a Staff
    employee's Full-OT days zeroed out of total_ot_hours — they worked a
    declared Holiday/Paid Holiday/Comp Off (see overtime_view's
    docstring), so that day converts to an EL day instead of paid OT
    (see LeaveLedgerEntry), rather than adding to OT hours/amount like
    every other subcategory's Full-OT day still does. Adds an "is_el_day"
    column so callers can count those days separately (e.g. the OT page's
    EL Days column) instead of just dropping them.

    Used everywhere OT hours/amount get aggregated per employee —
    _build_month_grid, and the OT page's Monthly Summary and department
    breakdown — so none of them can drift from which days actually still
    count as OT for Staff. The day-cell-level hover tooltip is the one
    exception: it's built from the raw, unadjusted shift_ot_table instead
    (see _build_month_grid's shift_ot_map), so it keeps showing what was
    actually worked that day regardless of how the monthly total treats
    it."""
    if shift_ot_table.empty:
        return shift_ot_table.assign(is_el_day=pd.Series(dtype=bool))
    table = shift_ot_table.copy()
    is_el_day = table["emp_code"].isin(staff_codes) & (table["full_day_ot"] == 1)
    table["is_el_day"] = is_el_day
    table["total_ot_hours"] = table["total_ot_hours"].where(~is_el_day, 0.0)
    return table


def _apply_staff_ot_display(table_rows: list, shift_ot_table: pd.DataFrame) -> None:
    """Makes a Staff day cell read exactly like any other employee's real
    M-OT/E-OT/ME-OT/Full-OT day: colors it by overtime_view's
    effective_shift (their actual AttendanceRecord.shift stays whatever
    was punched, typically "GS", so _build_month_grid never sets
    shift_flag for them on its own) and, on a Full-OT day, sets
    el_day_credit so the template can show "+1 EL"/"+0.5 EL" instead of
    OT hours — that day converts to EL, not paid OT (see
    _ot_payable_table).

    Mutates table_rows in place. Called on the OT page's grid["table_rows"]
    itself (see _ot_details_context) before it's deep-copied for the
    editable OT View tab and separately restricted/filtered for the
    read-only Full Monthly View tab, so a Staff employee's EL-earning day
    reads the same "+EL" way on both — distinct from the Attendance
    dashboard, which never calls this and keeps Staff on its own
    hours-heat-map treatment."""
    if shift_ot_table.empty:
        return
    effective_shift_map = {
        (r.emp_code, r.date): r.effective_shift for r in shift_ot_table.itertuples(index=False)
    }
    el_day_credit_map = {
        (r.emp_code, r.date): r.el_day_credit for r in shift_ot_table.itertuples(index=False)
    }
    for row in table_rows:
        if row["is_dept"]:
            continue
        for cell in row["day_cells"]:
            if not cell["is_staff"] or not cell["date_iso"]:
                continue
            key = (cell["emp_code"], pd.Timestamp(cell["date_iso"]))
            effective_shift = effective_shift_map.get(key, "")
            if effective_shift not in metrics.OT_SHIFT_CODES:
                continue
            cell["shift_flag"] = effective_shift
            cell["el_day_credit"] = el_day_credit_map.get(key, 0.0)


def _merge_edited_fields(existing: str, changed: list[str]) -> str:
    """Union `changed` field labels ("Punch In"/"Punch Out"/"Shift"/
    "Status") into the existing comma-separated manually_edited_fields
    value, preserving first-seen order and never dropping a field a
    previous edit already recorded, even if this edit didn't touch it."""
    fields = [f.strip() for f in existing.split(",") if f.strip()]
    for field in changed:
        if field not in fields:
            fields.append(field)
    return ", ".join(fields)


def _append_edited_fields_line(title: str, edited_fields: str) -> str:
    """Appends an "EDITED - ..." line to a day-cell's punch tooltip (see
    metrics.punch_time_labels) naming which field(s) a hand edit actually
    changed, so the tooltip says more than just "this was edited"."""
    if not edited_fields:
        return title
    line = f"EDITED - {edited_fields}"
    return f"{title}\n{line}" if title else line


def _build_month_grid(
    daily: pd.DataFrame, special_days: dict, emp_rate_map: dict, ot_tooltip: bool = False,
    full_day_map: dict | None = None,
) -> dict:
    """Builds the day x employee grid (day_headers + table_rows, matching
    the Month_Attendance pivot layout) shared by the dashboard's Month
    Attendance grid and the OT page's tabs. Also returns a few
    intermediate values (working_days, issues, emp_ot_totals,
    emp_el_days) that callers still need for their own KPI/department
    cards, so they don't have to recompute them.

    full_day_map (date -> hours, from EarlyClosureDay — see
    _early_closure_hours) lowers the expected full day below the
    standard 8.5h for Short Days/Permission Hours on specific dates."""
    full_day_map = full_day_map or {}
    working_days = metrics.infer_working_days(daily, special_days)

    # Shift-based OT (M-OT/E-OT/ME-OT/Full-OT), same calculation as the OT
    # Details report — drives every OT figure here except the day-cell's
    # bold-red OT text/background, which stays tied to special-day-worked
    # hours.
    shift_ot_table = metrics.overtime_view(daily)
    # The day-cell hover tooltip uses the raw table (see time_labels
    # below) — it should keep showing what was actually worked that day
    # regardless of how the monthly total treats it.
    shift_ot_map = (
        {(r.emp_code, r.date): r.total_ot_hours for r in shift_ot_table.itertuples(index=False)}
        if not shift_ot_table.empty else {}
    )
    staff_codes = (
        set(daily.loc[daily["subcategory"] == "Staff", "emp_code"])
        if "subcategory" in daily.columns else set()
    )
    ot_payable_table = _ot_payable_table(shift_ot_table, staff_codes)
    emp_ot_totals = (
        ot_payable_table.groupby("emp_code")["total_ot_hours"].sum().to_dict()
        if not ot_payable_table.empty else {}
    )
    emp_el_days = (
        ot_payable_table[ot_payable_table["is_el_day"]].groupby("emp_code")["el_day_credit"].sum().to_dict()
        if not ot_payable_table.empty else {}
    )
    emp_permission_hours = metrics.permission_hours_by_employee(daily, full_day_map)

    first_date = pd.Timestamp(daily["date"].iloc[0])
    _, days_in_month = py_calendar.monthrange(first_date.year, first_date.month)
    dates = [
        pd.Timestamp(year=first_date.year, month=first_date.month, day=d)
        for d in range(1, days_in_month + 1)
    ]
    month_view, day_labels = metrics.month_attendance_view(
        daily, working_days, dates=dates, full_day_map=full_day_map,
        department_order=settings.ATTENDANCE_VISIBLE_DEPARTMENTS,
    )
    time_labels = metrics.punch_time_labels(
        daily, shift_ot_map, ot_rate_map=emp_rate_map if ot_tooltip else None
    )
    issues = metrics.punch_issues(daily)
    special_status_map = {(r.emp_code, r.date): r.status for r in daily.itertuples(index=False)}
    ot_map = {(r.emp_code, r.date): r.ot_hours for r in daily.itertuples(index=False)}
    punch_map = {
        (r.emp_code, r.date): (r.shift, r.time_in, r.time_out) for r in daily.itertuples(index=False)
    }
    staff_map = {
        (r.emp_code, r.date): r.subcategory == "Staff" for r in daily.itertuples(index=False)
    }
    special_worked_map = {
        (r.emp_code, r.date): r.special_worked for r in daily.itertuples(index=False)
    }
    manually_edited_map = {
        (r.emp_code, r.date): r.manually_edited for r in daily.itertuples(index=False)
    }
    manually_edited_fields_map = {
        (r.emp_code, r.date): r.manually_edited_fields for r in daily.itertuples(index=False)
    }
    dept_map = {
        (r.emp_code, r.date): r.department for r in daily.itertuples(index=False)
    }
    special_day_codes = {SpecialDay.HOLIDAY, SpecialDay.PAID_HOLIDAY, SpecialDay.COMP_OFF}
    day_headers = [
        {
            "label": d,
            "dow": _DOW_LABELS[pd.Timestamp(date).dayofweek],
            "special": special_days.get(pd.Timestamp(date).date(), ""),
            "date_iso": pd.Timestamp(date).strftime("%Y-%m-%d"),
        }
        for d, date in zip(day_labels, dates)
    ]
    # Must exactly match metrics.month_attendance_view's own summary_cols
    # (in the same order) — month_view's columns come straight from that
    # function, and table_row construction below does row[c] for c in
    # summary_cols, so a mismatch here would KeyError.
    summary_cols = [
        "Work Days", "Comp Off", "EL", "Paid Holiday", "Personal Leave",
        "Missing Punch", "Short Days", "Permission Hours",
    ]

    table_rows = []
    for _, row in month_view.iterrows():
        is_dept = row["Row Labels"].startswith("▸")
        day_cells = []
        for d, date in zip(day_labels, dates):
            key = (row["Emp Code"], date)
            day_full_hours = full_day_map.get(pd.Timestamp(date).date(), 8.5)
            emp_status = special_status_map.get(key)
            # Holiday/Paid Holiday/Comp Off cells always keep the special
            # background, whether or not the employee worked that day — an
            # hours-based heat-map color doesn't apply on OT days, so
            # special_worked only drives the bold-red OT text (see "ot"
            # below), not the background.
            is_special_cell = not is_dept and emp_status in special_day_codes
            edited_fields_set = (
                set() if is_dept
                else {f.strip() for f in manually_edited_fields_map.get(key, "").split(",") if f.strip()}
            )
            shift, time_in, time_out = punch_map.get(key, ("", "", ""))
            time_in = metrics.clean_punch_time(time_in)
            time_out = metrics.clean_punch_time(time_out)
            # Gated on "not already a confirmed OT shift" rather than
            # shift == "GS" specifically — a blank/missing shift value
            # (some rows never got a code at all) is just as eligible for
            # a suggestion as an explicit "GS", and excluding it here was
            # silently hiding real OT-eligible days that had no shift
            # code of any kind.
            suggested_shift = (
                metrics.suggest_shift_code(
                    time_in, time_out, special_worked_map.get(key, False),
                    dept_map.get(key, ""), row[d],
                )
                if not is_dept and shift not in metrics.OT_SHIFT_CODES and not staff_map.get(key, False) else ""
            )
            day_cells.append({
                "value": "" if pd.isna(row[d]) else f"{row[d]:.1f}",
                "band": metrics.hours_heat_band(row[d]),
                # A person corrected this cell by hand (edit_record_view
                # or bulk_set_shift_view — see AttendanceRecord.
                # manually_edited), as opposed to it just reflecting
                # whatever the device export said — flagged on Device
                # Records so HR can see at a glance which cells aren't
                # raw device data anymore.
                "manually_edited": not is_dept and bool(manually_edited_map.get(key, False)),
                # Which specific field(s) were hand-edited (see
                # AttendanceRecord.manually_edited_fields) — each drives
                # its own differently-colored corner-dot marker in
                # dashboard.html (.edited-time/.edited-shift/.edited-status)
                # instead of one generic "this was edited" flag, so HR can
                # tell at a glance what kind of correction it was without
                # needing a separate audit-log page.
                "edited_time": bool({"Punch In", "Punch Out"} & edited_fields_set),
                "edited_shift": "Shift" in edited_fields_set,
                "edited_status": "Status" in edited_fields_set,
                # Edited before manually_edited_fields existed, so which
                # field changed was never captured — a neutral grey dot
                # instead of one of the three specific colors above (or
                # nothing at all, which would make an old edit silently
                # stop showing as edited once this field-level tracking
                # shipped).
                "edited_unknown": (
                    not is_dept and bool(manually_edited_map.get(key, False)) and not edited_fields_set
                ),
                # Text-color-only flag (see dashboard.html's .short-hours
                # rule) for a real day worked short of the usual 8.5h full
                # day — not on a special-day cell, which already has its
                # own dedicated color story (and is excluded from the
                # Short Days summary count for the same reason).
                "is_short": (
                    not is_dept and not is_special_cell and metrics.is_short_hours(row[d], day_full_hours)
                ),
                # The shortfall itself (that date's expected full day
                # minus actual hours — see day_full_hours/EarlyClosureDay
                # above) — shown instead of the raw worked hours on the
                # OT page's "-" short-day cells (see ot_details.html),
                # same figure Permission Hours sums across the month.
                # Rendered as a compact "Xh Ym"/"Ym" label, not decimal
                # hours.
                "short_hours_label": (
                    metrics.format_hours_as_hm(day_full_hours - row[d])
                    if not is_dept and not is_special_cell and metrics.is_short_hours(row[d], day_full_hours)
                    else ""
                ),
                "title": (
                    ""
                    if is_dept
                    else _append_edited_fields_line(
                        time_labels.get(key, ""), manually_edited_fields_map.get(key, "")
                    )
                ),
                "issue": not is_dept and key in issues,
                "special": emp_status if is_special_cell else "",
                "leave": not is_dept and emp_status in ("A", "PL"),
                "ot": not is_dept and ot_map.get(key, 0) > 0,
                "emp_code": "" if is_dept else row["Emp Code"],
                "date_iso": "" if is_dept else pd.Timestamp(date).strftime("%Y-%m-%d"),
                "status": "" if is_dept else (emp_status or ""),
                "shift": "" if is_dept else shift,
                "time_in": "" if is_dept else time_in,
                "time_out": "" if is_dept else time_out,
                "is_staff": not is_dept and staff_map.get(key, False),
                "shift_flag": shift if (not is_dept and shift in metrics.OT_SHIFT_CODES) else "",
                # Overridden per-cell (Staff, OT View tab only) by
                # _apply_staff_ot_display — see there for why this stays 0
                # here rather than being computed for every employee.
                "el_day_credit": 0,
                "shift_ot_hours": (
                    "" if is_dept or not shift_ot_map.get(key)
                    else f"{shift_ot_map[key]:.2f}".rstrip("0").rstrip(".")
                ),
                "suggested_shift": suggested_shift,
                # Staff always show their punch time here (like HOUSE
                # KEEPING) rather than needing a suggestion first — since
                # metrics.overtime_view now treats every Staff day as
                # "ME-OT" automatically (see its docstring), there's no
                # shift-code-assignment step for HR to be nudged into in
                # the first place, so the day just needs to be visible at
                # all for its already-computed OT (shift_ot_hours above)
                # to actually show up in this grid.
                "gs_show_time": (
                    not is_dept and shift not in metrics.OT_SHIFT_CODES and bool(time_in) and bool(time_out)
                    and (
                        staff_map.get(key, False)
                        or bool(suggested_shift)
                        or dept_map.get(key, "") == "HOUSE KEEPING"
                    )
                ),
            })
        summary_cells = [row[c] for c in summary_cols]
        total_ot = emp_ot_totals.get(row["Emp Code"], 0) if not is_dept else 0
        ot_rate = float(emp_rate_map.get(row["Emp Code"], 0)) if not is_dept else 0
        total_ot_amount = float(total_ot) * ot_rate if not is_dept else 0
        el_days = emp_el_days.get(row["Emp Code"], 0) if not is_dept else 0
        permission_hours = emp_permission_hours.get(row["Emp Code"], 0) if not is_dept else 0
        table_rows.append({
            "label": row["Row Labels"],
            "is_dept": is_dept,
            "day_cells": day_cells,
            "summary_cells": summary_cells,
            "total_ot": "" if is_dept or not total_ot else round(float(total_ot), 2),
            # Display-only "Xh Ym" label for the Total OT column (see
            # format_hours_as_hm) — "total_ot" itself stays a plain
            # decimal number, since _ot_only_rows/Excel export/etc. still
            # need to do real arithmetic/truthiness checks on it.
            "total_ot_label": "" if is_dept or not total_ot else metrics.format_hours_as_hm(total_ot),
            "ot_rate": "" if is_dept or not ot_rate else round(ot_rate, 2),
            "total_ot_amount": "" if is_dept or not total_ot_amount else round(total_ot_amount, 2),
            "el_days": "" if is_dept or not el_days else el_days,
            "permission_hours": "" if is_dept else metrics.format_hours_as_hm(permission_hours),
            # Total OT minus Permission Hours — the actual OT owed once
            # that month's short-day shortfall is netted against it.
            # Genuinely allowed to go negative (they owe more time than
            # they earned in OT) rather than floored at 0, since that's
            # real information HR needs to see, not an error state.
            "paid_ot_hours": (
                "" if is_dept
                else metrics.format_hours_as_hm(float(total_ot) - float(permission_hours), allow_negative=True)
            ),
        })

    return {
        "working_days": working_days,
        "shift_ot_table": shift_ot_table,
        "emp_ot_totals": emp_ot_totals,
        "emp_el_days": emp_el_days,
        "emp_permission_hours": emp_permission_hours,
        "issues": issues,
        "day_labels": day_labels,
        "day_headers": day_headers,
        "summary_cols": summary_cols,
        "table_rows": table_rows,
        "total_cols": 1 + len(day_labels) + len(summary_cols) + 3,
    }


def _ot_only_rows(table_rows: list) -> list:
    """Keeps only employee rows with actual OT and/or EL days that month
    (a Staff employee whose only qualifying day this month was a
    Full-OT/EL one has total_ot == 0 — see _ot_payable_table — so el_days
    has to be checked too, or they'd vanish from this list entirely), or
    at least one unconfirmed worked-holiday day (see
    _is_unconfirmed_special_worked — this one earns zero real OT/EL yet,
    but still needs to survive to this list so that day stays visible on
    the report instead of dropping the whole employee), and drops any
    department header left with no employees under it afterward — used
    by the OT Details report's Full Monthly View tab, which should only
    ever list people who had OT, EL, or a still-unconfirmed worked
    holiday that month, not the whole roster with everything but OT
    cells blanked out."""
    filtered = []
    pending_dept = None
    pending_emps: list = []

    def flush():
        if pending_dept is not None and pending_emps:
            filtered.append(pending_dept)
            filtered.extend(pending_emps)

    for row in table_rows:
        if row["is_dept"]:
            flush()
            pending_dept = row
            pending_emps = []
        elif (
            row.get("total_ot") or row.get("el_days")
            or any(_is_unconfirmed_special_worked(c) for c in row["day_cells"])
        ):
            pending_emps.append(row)
    flush()
    return filtered


def _is_unconfirmed_special_worked(cell: dict) -> bool:
    """True for a day someone actually worked through a company Holiday/
    Paid Holiday/Comp Off (real punches present) but nobody has confirmed
    a Full-OT shift code for yet, so it currently carries zero real OT
    credit — see _restrict_to_ot_cells for why this one case is kept
    visible there despite that, unlike an ordinary unconfirmed M-OT/
    E-OT/ME-OT suggestion (a few minutes either side of normal hours,
    low-stakes enough to stay hidden until confirmed). A whole
    unconfirmed holiday worked is consequential — HR needs to be able to
    spot it on the "official" Full Monthly View report, not just stumble
    onto it in the editable OT View tab."""
    return bool(cell["special"] and cell["time_in"] and cell["time_out"] and not cell["shift_ot_hours"])


def _restrict_to_ot_cells(table_rows: list) -> list:
    """Empties every day cell that didn't actually earn real OT credit —
    used alongside _ot_only_rows by the OT Details report's Full Monthly
    View, so it only ever shows numbers that match overtime_view()'s real
    per-day totals, not the dashboard's "this looks like it should have
    had a shift code" suggestion nudge (which is a workflow aid for fixing
    data, not an OT amount — a suggested day with no shift code actually
    set carries zero real credit, same as an assigned shift that ended up
    earning nothing, e.g. clocked out just before the 5:30pm OT cutoff) —
    except an unconfirmed worked-holiday day (see
    _is_unconfirmed_special_worked), which stays visible (background,
    suggested-shift marker, punch time) purely for visibility; it still
    contributes nothing to any OT hours/amount total, since
    cell["shift_ot_hours"] (this cell's real credit) is left untouched
    either way. cell["shift_ot_hours"] is the single source of truth for
    whether a cell counts as real OT here.

    Mutates and returns table_rows in place; the dashboard's own grid is
    unaffected since it builds its own separate copy per request."""
    for row in table_rows:
        if row["is_dept"]:
            continue
        for cell in row["day_cells"]:
            has_credit = bool(cell["shift_ot_hours"])
            unconfirmed_special = _is_unconfirmed_special_worked(cell)
            if not has_credit and not unconfirmed_special:
                cell["value"] = ""
                cell["shift_flag"] = ""
                cell["gs_show_time"] = False
                cell["suggested_shift"] = ""
            elif unconfirmed_special:
                cell["gs_show_time"] = True
            else:
                cell["gs_show_time"] = not cell["shift_flag"]
                cell["suggested_shift"] = ""
    return table_rows


def _attendance_visible_mask(daily: pd.DataFrame, departments: list[str]):
    """Boolean mask selecting rows whose department is in `departments`
    (case-insensitive) — None if `departments` is empty, meaning "no
    restriction, keep every row" (dashboard_view treats None that way).
    See settings.ATTENDANCE_VISIBLE_DEPARTMENTS."""
    if not departments:
        return None
    wanted = {d.lower() for d in departments}
    return daily["department"].str.lower().isin(wanted)


@login_required
def dashboard_view(request):
    """Server-rendered Month Attendance dashboard: KPIs, department
    breakdown, holiday calendar, and the day x employee hours grid.

    Scoped to one calendar month at a time (with prev/next nav, like
    ot_view/calendar_view) — otherwise the grid mixes every date ever
    uploaded into one non-continuous table. Nav is driven by a single
    "date" query param (any day within the target month)."""
    daily_all = _load_daily_data()
    month_keys_all = (
        sorted({(ts.year, ts.month) for ts in daily_all["date"]}) if not daily_all.empty else []
    )
    if not month_keys_all:
        return render(request, "attendance/dashboard.html", {"empty": True, "has_data": False})

    default_date = date_cls(*month_keys_all[-1], 1)
    date_param = request.GET.get("date")
    current = _parse_month_date(date_param, default_date)
    year, month = current.year, current.month

    prev_date = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_date = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
    current_view = request.GET.get("view", "all")
    if current_view not in ("all", "issues"):
        current_view = "all"
    locked_views = set(
        MonthLock.objects.filter(year=year, month=month).values_list("view", flat=True)
    )
    month_nav = {
        "has_data": True,
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
        "current_date": current.isoformat(),
        "current_view": current_view,
        "current_view_label": dict(MonthLock.VIEW_CHOICES)[current_view],
        "is_locked": current_view in locked_views,
        "lock_status": {v: (v in locked_views) for v, _ in MonthLock.VIEW_CHOICES if v != MonthLock.VIEW_OT},
    }

    daily = daily_all[(daily_all["date"].dt.year == year) & (daily_all["date"].dt.month == month)]
    visible_mask = _attendance_visible_mask(daily, settings.ATTENDANCE_VISIBLE_DEPARTMENTS)
    if visible_mask is not None:
        daily = daily[visible_mask]
    if daily.empty:
        return render(request, "attendance/dashboard.html", {"empty": True, **month_nav})

    special_days, downgraded_special_days, skip_special_days = _special_days_and_downgrades()
    daily = metrics.apply_special_days(daily, special_days, downgraded_special_days, skip_special_days)

    holidays = metrics.holiday_dates(daily)
    paid_holidays = metrics.paid_holiday_dates(daily)
    comp_offs = metrics.comp_off_dates(daily)

    emp_rate_map = dict(Employee.objects.values_list("code", "ot_rate_per_hour"))
    full_day_map = _early_closure_hours()
    grid = _build_month_grid(daily, special_days, emp_rate_map, full_day_map=full_day_map)
    working_days = grid["working_days"]
    issues = grid["issues"]

    emp_summary = metrics.employee_summary(daily, working_days, shift_ot_map=grid["emp_ot_totals"])
    dept_summary = metrics.department_summary(emp_summary)
    kpi = metrics.kpis(emp_summary)

    extremes = metrics.department_attendance_extremes(emp_summary, n=2)
    high_leave = metrics.department_high_leave(emp_summary, threshold=4)
    dept_summary_rows = dept_summary.to_dict("records")
    for d in dept_summary_rows:
        d.update(extremes.get(d["department"], {"top": [], "worst": []}))
        d["high_leave"] = high_leave.get(d["department"], [])

    dept_missed_punch = metrics.department_missed_punch_summary(daily, issues)

    context = {
        **month_nav,
        "empty": False,
        "period_start": daily["date"].min(),
        "period_end": daily["date"].max(),
        "working_days": working_days,
        "holidays": holidays,
        "paid_holidays": paid_holidays,
        "comp_offs": comp_offs,
        "kpi": kpi,
        "missed_punch_count": len(issues),
        "dept_summary": dept_summary_rows,
        "dept_missed_punch": dept_missed_punch,
        "day_labels": grid["day_labels"],
        "day_headers": grid["day_headers"],
        "summary_cols": grid["summary_cols"],
        "table_rows": grid["table_rows"],
        "heat_colors": metrics.HEAT_COLORS,
        # Not grid["total_cols"] — that includes the 3 OT columns, which
        # only the OT page's grid renders now that OT View has moved
        # there (see _ot_details_context).
        "total_cols": 1 + len(grid["day_headers"]) + len(grid["summary_cols"]),
        "day_types": SpecialDay.TYPE_CHOICES,
        "status_choices": AttendanceRecord.STATUS_CHOICES,
    }
    return render(request, "attendance/dashboard.html", context)


def _build_month_weeks(year, month, special_map, today, early_map: dict | None = None):
    """Monday-start week grid for one month, each day annotated with its
    SpecialDay type (if any) and EarlyClosureDay closing time (if any —
    see early_map, date -> EarlyClosureDay). Shared by calendar_view
    (editable) and dashboard_view (read-only preview of the period being
    viewed)."""
    early_map = early_map or {}
    cal = py_calendar.Calendar(firstweekday=0)
    month_weeks = cal.monthdatescalendar(year, month)
    return [
        [
            {
                "date": d,
                "in_month": d.month == month,
                "is_today": d == today,
                "day_type": special_map[d].day_type if d in special_map else "",
                "name": special_map[d].name if d in special_map else "",
                "early_closure_time": early_map[d].closing_time if d in early_map else None,
            }
            for d in week
        ]
        for week in month_weeks
    ]


@login_required
def calendar_view(request):
    """Company-wide Holiday / Paid Holiday / Comp Off calendar. Click a day
    to set/clear its type — each change is its own POST + redirect back to
    the same month, so no JS beyond auto-submitting the select is needed.
    Each day also has an optional Early Closure time field (see
    EarlyClosureDay) — a date the whole company closes earlier than
    usual, lowering the expected full day (computed from the standard
    9:00 AM start) for Short Days/Permission Hours company-wide on that
    date instead of every early leaver getting flagged/docked for it."""
    if request.method == "POST":
        action = request.POST.get("action", "")
        day_str = request.POST.get("date")

        if action == "early_closure":
            time_str = request.POST.get("closing_time", "").strip()
            if day_str:
                if time_str:
                    closing_time = parse_time(time_str)
                    if closing_time is None:
                        _error(request, "Enter a valid closing time.")
                        return redirect(
                            f"{request.path}?year={request.POST.get('year')}&month={request.POST.get('month')}"
                        )
                    EarlyClosureDay.objects.update_or_create(date=day_str, defaults={"closing_time": closing_time})
                else:
                    EarlyClosureDay.objects.filter(date=day_str).delete()
            return redirect(
                f"{request.path}?year={request.POST.get('year')}&month={request.POST.get('month')}"
            )

        day_type = request.POST.get("day_type", "")
        if day_str:
            if day_type in dict(SpecialDay.TYPE_CHOICES):
                SpecialDay.objects.update_or_create(date=day_str, defaults={"day_type": day_type})
            else:
                SpecialDay.objects.filter(date=day_str).delete()
        return redirect(
            f"{request.path}?year={request.POST.get('year')}&month={request.POST.get('month')}"
        )

    today = date_cls.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    cal = py_calendar.Calendar(firstweekday=0)  # Monday-start weeks
    month_weeks = cal.monthdatescalendar(year, month)
    visible_dates = [d for week in month_weeks for d in week]
    special_map = {sd.date: sd for sd in SpecialDay.objects.filter(date__in=visible_dates)}
    early_map = {ec.date: ec for ec in EarlyClosureDay.objects.filter(date__in=visible_dates)}

    weeks = _build_month_weeks(year, month, special_map, today, early_map)

    prev_month, prev_year = (12, year - 1) if month == 1 else (month - 1, year)
    next_month, next_year = (1, year + 1) if month == 12 else (month + 1, year)

    context = {
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "weeks": weeks,
        "day_types": SpecialDay.TYPE_CHOICES,
        "prev_year": prev_year, "prev_month": prev_month,
        "next_year": next_year, "next_month": next_month,
        "upcoming": SpecialDay.objects.filter(date__gte=today).order_by("date")[:20],
    }
    return render(request, "attendance/calendar.html", context)


def _month_is_locked(date_str: str, view: str) -> bool:
    try:
        d = date_cls.fromisoformat(date_str)
    except ValueError:
        return False
    return MonthLock.objects.filter(year=d.year, month=d.month, view=view).exists()


@login_required
def edit_record_view(request):
    """HR correction for a single employee/date cell — fixes a missed punch
    (time_in/time_out), sets the shift code (GS/M-OT/E-OT/ME-OT/Full-OT)
    so it feeds correctly into the OT view, and/or forces the day's
    status (e.g. Earned Leave, Personal Leave, Comp Off) instead of
    letting it auto-derive from the punch times. Triggered by the
    click-to-edit popup on either the Attendance dashboard's grid or the
    OT page's OT View tab; redirects back to wherever the popup was
    opened from."""
    next_url = request.POST.get("next") or "dashboard"
    if request.method != "POST":
        return redirect(next_url)

    emp_code = request.POST.get("emp_code", "").strip()
    date_str = request.POST.get("date", "").strip()
    time_in = request.POST.get("time_in", "").strip()
    time_out = request.POST.get("time_out", "").strip()
    shift = request.POST.get("shift", "").strip()
    status_override = request.POST.get("status", "").strip()
    view = request.POST.get("view", "all")

    if _month_is_locked(date_str, view):
        _error(request, "This view is locked for this month — unlock it first to make changes.")
        return redirect(next_url)

    try:
        record = AttendanceRecord.objects.get(employee__code=emp_code, date=date_str)
    except AttendanceRecord.DoesNotExist:
        _error(request, f"No attendance record for {emp_code} on {date_str}.")
        return redirect(next_url)

    old_time_in, old_time_out = record.time_in, record.time_out
    old_shift, old_status = record.shift, record.status

    record.time_in = time_in
    record.time_out = time_out
    # Staff are excluded from shift-based OT entirely (see overtime_view) —
    # their shift is never editable, regardless of what the form submits,
    # so a disabled/bypassed field can't silently change or clear it.
    if record.employee.subcategory != "Staff":
        record.shift = shift
    record.work_hours, computed_status = metrics.recompute_from_punch(time_in, time_out)
    # An explicit status (e.g. Earned Leave — see AttendanceRecord.
    # STATUS_CHOICES) always wins over the auto-derived P/HD/A from the
    # punch times; leaving the picker on its blank "Auto" option (the
    # default whenever the day's current status is itself one of P/HD/A,
    # rather than an already-forced one — see the popup's own JS) keeps
    # today's behavior unchanged.
    record.status = status_override if status_override in dict(AttendanceRecord.STATUS_CHOICES) else computed_status
    record.manually_edited = True

    changed_fields = []
    if record.time_in != old_time_in:
        changed_fields.append("Punch In")
    if record.time_out != old_time_out:
        changed_fields.append("Punch Out")
    if record.shift != old_shift:
        changed_fields.append("Shift")
    if record.status != old_status:
        changed_fields.append("Status")
    record.manually_edited_fields = _merge_edited_fields(record.manually_edited_fields, changed_fields)

    record.save()
    # On a date that's a company-wide Paid Holiday/Comp Off, apply_special_days
    # (see src/metrics.py) unconditionally overwrites status to that day_type
    # for everyone on every read — so picking "Holiday" here wouldn't actually
    # stick on the next page load unless this employee is registered as
    # downgraded for that date too. Picking anything else (including "Auto")
    # un-registers them, so the calendar's real day_type resumes applying —
    # this is the one place that toggle is meant to be flipped from (see
    # SpecialDay.downgraded_employees).
    special_day = SpecialDay.objects.filter(date=date_str).exclude(day_type=SpecialDay.HOLIDAY).first()
    if special_day:
        if record.status == SpecialDay.HOLIDAY:
            special_day.downgraded_employees.add(record.employee)
        else:
            special_day.downgraded_employees.remove(record.employee)
    messages.success(request, f"Updated {emp_code} on {date_str}.")
    logger.info(
        "Record edited: emp=%s date=%s time_in=%s time_out=%s shift=%s status=%s by user=%s",
        emp_code, date_str, time_in, time_out, shift, record.status, request.user,
    )
    return redirect(next_url)


@login_required
def bulk_set_shift_view(request):
    """Sets the shift code for every employee on one date at once, and/or
    the company-wide calendar day type (Holiday/Paid Holiday/Comp Off) for
    that same date — one popup, triggered by clicking a day column header
    in the Attendance dashboard's grid or the OT page's OT View tab,
    instead of separate shift and calendar-editing flows. Staff
    subcategory employees are excluded from the shift update,
    matching overtime_view's own exclusion (shift-based OT never applies
    to them); the calendar day type still applies company-wide."""
    next_url = request.POST.get("next") or "dashboard"
    if request.method != "POST":
        return redirect(next_url)

    date_str = request.POST.get("date", "").strip()
    shift = request.POST.get("shift", "").strip()
    day_type = request.POST.get("day_type", "")
    view = request.POST.get("view", "all")
    if not date_str:
        _error(request, "Missing date.")
        return redirect(next_url)

    if _month_is_locked(date_str, view):
        _error(request, "This view is locked for this month — unlock it first to make changes.")
        return redirect(next_url)

    records = list(
        AttendanceRecord.objects.filter(date=date_str).exclude(employee__subcategory="Staff")
    )
    for record in records:
        if record.shift != shift:
            record.manually_edited_fields = _merge_edited_fields(record.manually_edited_fields, ["Shift"])
        record.shift = shift
        record.manually_edited = True
    AttendanceRecord.objects.bulk_update(records, ["shift", "manually_edited", "manually_edited_fields"])
    updated = len(records)
    if day_type in dict(SpecialDay.TYPE_CHOICES):
        SpecialDay.objects.update_or_create(date=date_str, defaults={"day_type": day_type})
    else:
        SpecialDay.objects.filter(date=date_str).delete()

    messages.success(
        request,
        f"Set shift '{shift}' for {updated} employee(s) and calendar type '{day_type or '—'}' on {date_str}.",
    )
    logger.info(
        "Bulk shift set: date=%s shift=%s day_type=%s updated=%s by user=%s",
        date_str, shift, day_type or "-", updated, request.user,
    )
    return redirect(next_url)


@login_required
def toggle_month_lock_view(request):
    """Locks or unlocks one calendar month's attendance/payroll data for
    one or more views at once — the Attendance dashboard's All/Missed
    Punch views, the OT page's OT View tab, or any of the Salary page's
    five tabs — gated by the shared PIN (see
    settings.ATTENDANCE_LOCK_PIN) — there's no user login in this app, so
    the PIN is the only check on either direction. Triggered by a single
    lock-toggle button next to the view/tab it applies to (one "view"
    field) or a bulk lock/unlock picker with several checked "view"
    checkboxes (e.g. Salary's "Lock tabs" picker) — request.POST.getlist
    handles both the same way."""
    next_url = request.POST.get("next") or "dashboard"
    if request.method != "POST":
        return redirect(next_url)

    try:
        year = int(request.POST.get("year", ""))
        month = int(request.POST.get("month", ""))
    except ValueError:
        _error(request, "Missing month.")
        return redirect(next_url)

    view_choices = dict(MonthLock.VIEW_CHOICES)
    views = [v for v in request.POST.getlist("view") if v in view_choices]
    if not views:
        _error(request, "Select at least one tab/view to lock or unlock.")
        return redirect(next_url)

    pin = request.POST.get("pin", "").strip()
    action = request.POST.get("action", "")
    if pin != settings.ATTENDANCE_LOCK_PIN:
        messages.error(request, "Incorrect PIN.")
        logger.warning(
            "Incorrect lock PIN attempt: view=%s %s-%s by user=%s", ",".join(views), year, month, request.user,
        )
        return redirect(next_url)

    labels = []
    for view in views:
        if action == "lock":
            MonthLock.objects.get_or_create(year=year, month=month, view=view)
        elif action == "unlock":
            MonthLock.objects.filter(year=year, month=month, view=view).delete()
        labels.append(view_choices[view])
        logger.info(
            "Month lock %s: view=%s %s-%s by user=%s", action, view, year, month, request.user,
        )

    if action in ("lock", "unlock"):
        verb = "locked" if action == "lock" else "unlocked"
        are_is = "is" if len(labels) == 1 else "are"
        messages.success(
            request,
            f"{', '.join(labels)} {are_is} now {verb} for {py_calendar.month_name[month]} {year}.",
        )
    return redirect(next_url)


def _notify_telegram(text: str, parse_mode: str | None = None, topic_id: str = "") -> tuple[bool, str]:
    """Posts a text message to the configured bot/chat (see
    settings.TELEGRAM_*) and the given topic (message_thread_id) — each
    feature area passes its own settings.TELEGRAM_TOPIC_ID_* so Attendance/
    OT/Cash Withdrawal land in separate topics of the same group. Never
    raises — returns (ok, error) so a caller that needs to know (the Day
    view's "Send Report" button) can surface a failure, while a best-effort
    caller (the bulk save notification) can just ignore the result."""
    token = settings.TELEGRAM_BOT_TOKEN
    chat_id = settings.TELEGRAM_CHAT_ID
    if not token or not chat_id:
        return False, "Telegram isn't configured (missing bot token or chat id)."
    payload = {"chat_id": chat_id, "text": text}
    if topic_id:
        payload["message_thread_id"] = topic_id
    if parse_mode:
        payload["parse_mode"] = parse_mode
    request_obj = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
    )
    try:
        urllib.request.urlopen(request_obj, timeout=5)
        logger.info("Telegram message sent.")
        return True, ""
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", "replace")
        logger.warning("Telegram notification failed: %s", detail)
        return False, detail
    except urllib.error.URLError as exc:
        logger.warning("Telegram notification failed: %s", exc)
        return False, str(exc.reason)
    except Exception as exc:  # noqa: BLE001 - a timeout/SSL/connection error
        # isn't a URLError/HTTPError in every case (e.g. a read timeout can
        # raise a bare TimeoutError) — catch-all so a send failure is always
        # logged and reported back, never silently swallowed.
        logger.exception("Telegram notification failed unexpectedly: %s", exc)
        return False, str(exc)


def _telegram_send_photo(photo_bytes: bytes, filename: str, caption: str = "", topic_id: str = "") -> tuple[bool, str]:
    """Posts an image to the configured Telegram bot/chat via sendPhoto,
    into the given topic (message_thread_id) — see _notify_telegram for
    why this takes an explicit topic_id instead of one global setting.
    Unlike _notify_telegram, this one reports back (ok, error) instead of
    swallowing failures — it's triggered by an explicit "Send Report"
    click, so the user needs to know if it didn't go through."""
    token = settings.TELEGRAM_BOT_TOKEN
    chat_id = settings.TELEGRAM_CHAT_ID
    if not token or not chat_id:
        return False, "Telegram isn't configured (missing bot token or chat id)."

    fields = {"chat_id": chat_id, "caption": caption[:1024]}
    if topic_id:
        fields["message_thread_id"] = topic_id

    boundary = uuid.uuid4().hex
    body = bytearray()
    for name, value in fields.items():
        if not value:
            continue
        body += (
            f'--{boundary}\r\nContent-Disposition: form-data; name="{name}"\r\n\r\n{value}\r\n'
        ).encode("utf-8")
    body += (
        f'--{boundary}\r\nContent-Disposition: form-data; name="photo"; filename="{filename}"\r\n'
        f"Content-Type: image/png\r\n\r\n"
    ).encode("utf-8")
    body += photo_bytes
    body += f"\r\n--{boundary}--\r\n".encode("utf-8")

    request_obj = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendPhoto",
        data=bytes(body),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    try:
        # 45s, not 20 — a scale:2 screenshot of the Full Monthly View's wide
        # day x employee grid can run several MB, and the OT Details "Send
        # to Telegram" button sends two of these back-to-back, so the
        # second upload has to wait on the first's full round trip too.
        urllib.request.urlopen(request_obj, timeout=45)
        logger.info("Telegram photo sent.")
        return True, ""
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", "replace")
        logger.warning("Telegram sendPhoto failed: %s", detail)
        return False, detail
    except urllib.error.URLError as exc:
        logger.warning("Telegram sendPhoto failed: %s", exc)
        return False, str(exc.reason)
    except Exception as exc:  # noqa: BLE001 - a timeout/SSL/connection error
        # isn't a URLError/HTTPError in every case (e.g. a read timeout can
        # raise a bare TimeoutError), and a 20MB+ photo makes that more
        # likely than for the plain-text sendMessage above — catch-all so a
        # send failure is always logged and reported back, never silent.
        logger.exception("Telegram sendPhoto failed unexpectedly: %s", exc)
        return False, str(exc)


def _pick_department(request, departments):
    """Shared default-department resolution for the Mark Attendance flow —
    the explicit request param wins, else "Operators" (this flow's main
    use case), else whatever department happens to exist first."""
    dept_id = request.GET.get("department") or request.POST.get("department")
    if dept_id:
        dept = Department.objects.filter(id=dept_id).first()
        if dept:
            return dept
    return Department.objects.filter(name__iexact="Operators").first() or (departments[0] if departments else None)


def _build_simple_month_grid(dept, year: int, month: int) -> dict:
    """Builds the employee x day-of-month grid (day_headers, table_rows,
    day_totals, total_present) behind Mark Attendance's Month view table —
    and, rendered off-screen, behind the Day view's "Send Report" button
    too, so it can screenshot a month table without navigating away. Plain
    DB reads keyed by (employee, day-of-month); unrelated to the
    dashboard's pandas-based _build_month_grid, which draws from a
    different data source (_load_daily_data)."""
    _, days_in_month = py_calendar.monthrange(year, month)
    day_headers = [
        {
            "day": d,
            "dow": _DOW_LABELS[date_cls(year, month, d).weekday()],
            "date_iso": date_cls(year, month, d).isoformat(),
        }
        for d in range(1, days_in_month + 1)
    ]

    employees = (
        Employee.objects.filter(department=dept)
        .active_during(date_cls(year, month, 1), date_cls(year, month, days_in_month))
        .order_by("name")
        if dept else Employee.objects.none()
    )
    status_map = {
        (r["employee_id"], r["date"].day): r["status"]
        for r in AttendanceRecord.objects.filter(
            employee__in=employees, date__year=year, date__month=month
        ).values("employee_id", "date", "status")
    }
    total_days = len(day_headers)
    day_totals = [0] * total_days
    table_rows = []
    for emp in employees:
        cells = [
            {
                "day": d["day"],
                "date_iso": d["date_iso"],
                "status": status_map.get((emp.id, d["day"]), ""),
            }
            for d in day_headers
        ]
        present_count = 0
        absent_count = 0
        for i, c in enumerate(cells):
            if c["status"] == "P":
                present_count += 1
                day_totals[i] += 1
            elif not c["status"]:
                absent_count += 1
        table_rows.append({
            "employee": emp,
            "cells": cells,
            "present_count": present_count,
            "total_days": total_days,
            "percent_absent": round(absent_count / total_days * 100) if total_days else 0,
        })

    return {
        "day_headers": day_headers,
        "table_rows": table_rows,
        "day_totals": day_totals,
        "total_present": sum(day_totals),
    }


@login_required
def mark_attendance_view(request):
    """Daily attendance marking for one department at a time (Operators by
    default) — a fast, mobile-friendly alternative to waiting on the eSSL
    fingerprint export, for departments where punches aren't captured that
    way. One page load = one department + one date; marking is a single
    bulk POST that upserts an AttendanceRecord per employee, same
    upsert-by-(employee, date) semantics as the fingerprint importer."""
    date_param = request.GET.get("date") or request.POST.get("date")
    try:
        target_date = date_cls.fromisoformat(date_param) if date_param else date_cls.today()
    except ValueError:
        target_date = date_cls.today()

    departments = list(Department.objects.all())
    dept = _pick_department(request, departments)

    redirect_url = f"{request.path}?date={target_date.isoformat()}" + (f"&department={dept.id}" if dept else "")

    if request.method == "POST":
        if dept is None:
            _error(request, "No department selected.")
            return redirect(redirect_url)
        if _month_is_locked(target_date.isoformat(), MonthLock.VIEW_ALL):
            _error(request, "This month is locked — unlock it on the dashboard first.")
            return redirect(redirect_url)

        employees = Employee.objects.filter(department=dept).active_on(target_date)
        valid_status = dict(AttendanceRecord.STATUS_CHOICES)
        saved = 0
        present_count = 0
        absent_count = 0
        for emp in employees:
            # The Present/Absent toggle (radio, name=status_<id>) and the
            # "More…" dropdown (select, name=status_more_<id>) are separate
            # fields — only one is ever meaningful at a time (the template's
            # JS keeps them mutually exclusive), so an unchecked radio group
            # falls back to whatever the select holds.
            status = request.POST.get(f"status_{emp.id}", "").strip()
            if not status:
                status = request.POST.get(f"status_more_{emp.id}", "").strip()
            if status not in valid_status:
                continue
            shift = request.POST.get(f"shift_{emp.id}", "").strip()
            AttendanceRecord.objects.update_or_create(
                employee=emp, date=target_date,
                defaults={"status": status, "shift": shift},
            )
            saved += 1
            if status == "P":
                present_count += 1
            elif status == "A":
                absent_count += 1
        messages.success(
            request, f"Marked attendance for {saved} employee(s) in {dept.name} on {target_date:%d %b %Y}."
        )
        logger.info(
            "Day attendance saved: dept=%s date=%s saved=%s present=%s absent=%s by user=%s",
            dept.name, target_date, saved, present_count, absent_count, request.user,
        )
        if saved:
            other_count = saved - present_count - absent_count
            breakdown = f"{present_count} Present, {absent_count} Absent"
            if other_count:
                breakdown += f", {other_count} Other"
            _notify_telegram(
                f"🗓 Attendance marked — {dept.name} — {target_date:%d %b %Y}\n{saved} employee(s): {breakdown}",
                topic_id=settings.TELEGRAM_TOPIC_ID_ATTENDANCE,
            )
        return redirect(redirect_url)

    employees = (
        Employee.objects.filter(department=dept).active_on(target_date).order_by("name")
        if dept else Employee.objects.none()
    )
    existing = {
        r.employee_id: r
        for r in AttendanceRecord.objects.filter(date=target_date, employee__in=employees)
    }
    rows = [
        {
            "employee": emp,
            "status": existing[emp.id].status if emp.id in existing else "A",
            "shift": existing[emp.id].shift if emp.id in existing else "",
        }
        for emp in employees
    ]

    # "Send Report" sends the day's attendance as text plus a screenshot of
    # the whole month — the month grid below is rendered off-screen purely
    # so that screenshot can be captured without navigating to Month view.
    month_grid = _build_simple_month_grid(dept, target_date.year, target_date.month) if dept else None

    context = {
        "target_date": target_date,
        "prev_date": (target_date - timedelta(days=1)).isoformat(),
        "next_date": (target_date + timedelta(days=1)).isoformat(),
        "departments": departments,
        "dept": dept,
        "rows": rows,
        "status_choices": AttendanceRecord.STATUS_CHOICES,
        "is_locked": _month_is_locked(target_date.isoformat(), MonthLock.VIEW_ALL),
        "month_name": py_calendar.month_name[target_date.month],
        "month_day_headers": month_grid["day_headers"] if month_grid else [],
        "month_table_rows": month_grid["table_rows"] if month_grid else [],
        "month_day_totals": month_grid["day_totals"] if month_grid else [],
        "month_total_present": month_grid["total_present"] if month_grid else 0,
    }
    return render(request, "attendance/mark_attendance.html", context)


@login_required
def log_client_error_view(request):
    """Client-side failures (e.g. html2canvas throwing before any Telegram
    request is even made — see the "Send Report" buttons) are invisible to
    the server by definition, so there's nothing here for the normal error
    logging to catch. This just gives the browser a way to hand that
    failure to the server log instead of it vanishing in the console."""
    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)
    message = request.POST.get("message", "")[:2000]
    context = request.POST.get("context", "")[:200]
    logger.warning("Client-side error [%s, user=%s]: %s", context or "unknown", request.user, message)
    return JsonResponse({"ok": True})


@login_required
def send_day_attendance_report_view(request):
    """Sends a simple text attendance table (name — status, one per line)
    for one department/date to Telegram — the Day view's "Send Report"
    button. Plain text rather than a screenshot: the emp-list is just a
    column of radio toggles, which doesn't capture well as an image, so
    this rebuilds the same data as a Telegram-native monospace table."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required."}, status=405)

    dept = Department.objects.filter(id=request.POST.get("department")).first()
    if not dept:
        return JsonResponse({"ok": False, "error": "Invalid department."}, status=400)
    try:
        target_date = date_cls.fromisoformat(request.POST.get("date", ""))
    except ValueError:
        return JsonResponse({"ok": False, "error": "Invalid date."}, status=400)

    employees = Employee.objects.filter(department=dept).active_on(target_date).order_by("name")
    if not employees:
        return JsonResponse({"ok": False, "error": "No employees in this department."}, status=400)

    status_map = {
        r.employee_id: r.status
        for r in AttendanceRecord.objects.filter(date=target_date, employee__in=employees)
    }
    status_labels = dict(AttendanceRecord.STATUS_CHOICES)
    rows = [(emp.name, status_labels.get(status_map.get(emp.id, "A"), "Absent")) for emp in employees]

    name_width = max(len(name) for name, _ in rows)
    table_text = "\n".join(f"{name.ljust(name_width)}  {status}" for name, status in rows)
    present = sum(1 for _, status in rows if status == "Present")
    absent = sum(1 for _, status in rows if status == "Absent")

    text = (
        f"<b>🗓 Day Attendance — {html.escape(dept.name)} — {target_date:%d %b %Y}</b>\n"
        f"{present} Present, {absent} Absent\n"
        f"<pre>{html.escape(table_text)}</pre>"
    )
    try:
        ok, error = _notify_telegram(text, parse_mode="HTML", topic_id=settings.TELEGRAM_TOPIC_ID_ATTENDANCE)
    except Exception:
        logger.exception("send_day_attendance_report_view failed unexpectedly.")
        return JsonResponse({"ok": False, "error": "Unexpected server error."}, status=500)
    if not ok:
        return JsonResponse({"ok": False, "error": error or "Telegram send failed."}, status=502)
    return JsonResponse({"ok": True})


@login_required
def mark_attendance_month_view(request):
    """Whole-month companion to mark_attendance_view — one department at a
    time, a grid of employee x day-of-month, each cell directly editable
    (click it, pick a status, save) via set_attendance_status_view below.
    Read-heavy by design: one query for the whole month's records rather
    than N+1 per cell."""
    date_param = request.GET.get("date")
    try:
        current = date_cls.fromisoformat(date_param) if date_param else date_cls.today()
    except ValueError:
        current = date_cls.today()
    year, month = current.year, current.month

    departments = list(Department.objects.all())
    dept = _pick_department(request, departments)

    prev_date = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_date = (current.replace(day=28) + timedelta(days=4)).replace(day=1)

    grid = _build_simple_month_grid(dept, year, month)

    context = {
        "current": current,
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
        "departments": departments,
        "dept": dept,
        "day_headers": grid["day_headers"],
        "table_rows": grid["table_rows"],
        "day_totals": grid["day_totals"],
        "total_present": grid["total_present"],
        "status_choices": AttendanceRecord.STATUS_CHOICES,
        "status_labels": dict(AttendanceRecord.STATUS_CHOICES),
        "is_locked": _month_is_locked(current.isoformat(), MonthLock.VIEW_ALL),
    }
    return render(request, "attendance/mark_attendance_month.html", context)


_TELEGRAM_PHOTO_TOPICS = {
    "attendance": lambda: settings.TELEGRAM_TOPIC_ID_ATTENDANCE,
    "ot_cash": lambda: settings.TELEGRAM_TOPIC_ID_OT_CASH,
    "salary": lambda: settings.TELEGRAM_TOPIC_ID_SALARY,
}


@login_required
def send_telegram_report_view(request):
    """Receives a screenshot (captured client-side via html2canvas — see
    mark_attendance_month.html's "Send Report" button, and the same
    endpoint reused by OT Details and Cash Withdrawal's "Send to
    Telegram") and forwards it to Telegram as a photo, in the topic named
    by the "topic" field (see _TELEGRAM_PHOTO_TOPICS — one shared endpoint
    serving three feature areas needs the caller to say which topic it
    belongs in). JSON in, JSON out: the button is a fetch() call, not a
    form submit, so it can re-enable itself and show an error inline
    instead of navigating away."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required."}, status=405)
    image = request.FILES.get("image")
    if not image:
        return JsonResponse({"ok": False, "error": "No image provided."}, status=400)
    caption = request.POST.get("caption", "").strip()
    topic_key = request.POST.get("topic", "attendance")
    get_topic_id = _TELEGRAM_PHOTO_TOPICS.get(topic_key, _TELEGRAM_PHOTO_TOPICS["attendance"])
    try:
        ok, error = _telegram_send_photo(image.read(), image.name or "report.png", caption, topic_id=get_topic_id())
    except Exception:
        logger.exception("send_telegram_report_view failed unexpectedly.")
        return JsonResponse({"ok": False, "error": "Unexpected server error."}, status=500)
    if not ok:
        return JsonResponse({"ok": False, "error": error or "Telegram send failed."}, status=502)
    return JsonResponse({"ok": True})


def _send_email_report(subject: str, body: str, attachments: list[tuple[str, bytes]]) -> tuple[bool, str]:
    """Sends one email via the SMTP settings in config/settings.py, with
    every (filename, png_bytes) pair in attachments attached to the same
    message — the "Email" report button's server-side half. Unlike
    Telegram's one-photo-per-tab approach (_telegram_send_photo), every
    selected tab's screenshot rides along as an attachment on a single
    message, since the sender is meant to write one covering note in
    the compose box rather than getting several separate emails.
    Recipients are a fixed list (settings.EMAIL_RECIPIENTS) rather than
    typed per-send — there's no per-employee routing need here, just a
    small fixed distribution list HR maintains directly in settings."""
    if not settings.EMAIL_HOST or not settings.EMAIL_HOST_USER:
        return False, "Email isn't configured (missing SMTP host/user in .env)."
    if not settings.EMAIL_RECIPIENTS:
        return False, "No recipients configured (see settings.EMAIL_RECIPIENTS)."
    try:
        message = EmailMessage(
            subject=subject, body=body,
            from_email=settings.DEFAULT_FROM_EMAIL, to=settings.EMAIL_RECIPIENTS,
        )
        for filename, content in attachments:
            message.attach(filename, content, "image/png")
        message.send(fail_silently=False)
        logger.info("Email report sent: %r to %s (%d attachment(s)).", subject, settings.EMAIL_RECIPIENTS, len(attachments))
        return True, ""
    except Exception as exc:
        logger.warning("Email report send failed: %s", exc)
        return False, str(exc)


@login_required
def send_email_report_view(request):
    """Receives one or more screenshots (captured client-side via
    html2canvas, same as send_telegram_report_view) plus a subject/body
    typed into the Email button's compose box, and emails them as one
    message with every screenshot attached — see _send_email_report.
    JSON in, JSON out, same reasoning as send_telegram_report_view."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required."}, status=405)
    images = request.FILES.getlist("images")
    if not images:
        return JsonResponse({"ok": False, "error": "No attachments provided."}, status=400)
    subject = request.POST.get("subject", "").strip() or "Payday Report"
    body = request.POST.get("body", "")
    attachments = [(image.name or f"report-{i}.png", image.read()) for i, image in enumerate(images)]
    try:
        ok, error = _send_email_report(subject, body, attachments)
    except Exception:
        logger.exception("send_email_report_view failed unexpectedly.")
        return JsonResponse({"ok": False, "error": "Unexpected server error."}, status=500)
    if not ok:
        return JsonResponse({"ok": False, "error": error or "Email send failed."}, status=502)
    return JsonResponse({"ok": True})


@login_required
def set_attendance_status_view(request):
    """Single-cell save for the month grid — each day cell is its own tiny
    form whose submit button already carries the *target* status (the
    opposite of whatever's currently shown, computed in the template), so
    one click toggles Present <-> Absent with no popup. Same
    upsert-by-(employee, date) as mark_attendance_view's bulk save, just one
    record at a time."""
    next_url = request.POST.get("next") or "mark_attendance_month"
    if request.method != "POST":
        return redirect(next_url)

    date_str = request.POST.get("date", "").strip()
    if _month_is_locked(date_str, MonthLock.VIEW_ALL):
        _error(request, "This month is locked — unlock it on the dashboard first.")
        return redirect(next_url)

    try:
        emp = Employee.objects.get(id=request.POST.get("employee_id", "").strip())
        target_date = date_cls.fromisoformat(date_str)
    except (Employee.DoesNotExist, ValueError):
        _error(request, "Invalid employee or date.")
        return redirect(next_url)

    status = request.POST.get("status", "").strip()
    valid_status = dict(AttendanceRecord.STATUS_CHOICES)
    if status not in valid_status:
        _error(request, "Invalid status.")
        return redirect(next_url)

    AttendanceRecord.objects.update_or_create(
        employee=emp, date=target_date, defaults={"status": status},
    )
    messages.success(request, f"{emp.name}: {valid_status[status]} on {target_date:%d %b}.")
    logger.info(
        "Month-grid cell set: emp=%s date=%s status=%s by user=%s",
        emp.code, target_date, status, request.user,
    )
    return redirect(next_url)


# (tab key, page label, Employee.subcategory value) — Operators aren't
# looped in here since their context key doesn't follow the "{tab_key}_
# rows" pattern (see salary_view), and they're filtered by
# category="Operator" instead of subcategory, unlike the other three.
_SALARY_SUBCATEGORY_TABS = [
    ("company", "Company Workers", "Company"),
    ("helper", "Helpers", "Helper"),
    ("staff", "Staff", "Staff"),
]

# (tabs= query value, sheet label, _salary_context rows key) for all five
# Salary tabs — shared by salary_download_view and salary_bank_download_view
# for their ?tabs=... tab-picker filtering, and by the tab-picker forms in
# salary.html (the checkbox "value"s there must match these keys).
_SALARY_TAB_KEYS = [
    ("company", "Company Workers", "company_rows"),
    ("helper", "Helpers", "helper_rows"),
    ("staff", "Staff", "staff_rows"),
    ("contractors", "Contractors", "contractor_rows"),
    ("operators", "Operators", "operator_rows"),
    ("fixed_payments", "Fixed Payments", "fixed_payment_rows"),
]

# tab_key -> MonthLock.view constant — each Salary tab locks/unlocks
# independently (e.g. Company Workers can be finalized while Operators is
# still being entered), same PIN-gated mechanism as the Attendance
# dashboard/OT page (see toggle_month_lock_view). Keys must match the
# hidden "tab" field each tab's save <form> posts (salary.html) exactly —
# "helpers" (plural), not "helper", since that's what the Helpers form
# actually sends; a prior mismatch here silently skipped the server-side
# lock check for that one tab (the UI's disabled fieldset still blocked
# normal use, but a direct POST wouldn't have been refused).
_SALARY_LOCK_VIEWS = {
    "company": MonthLock.VIEW_SALARY_COMPANY,
    "helpers": MonthLock.VIEW_SALARY_HELPER,
    "staff": MonthLock.VIEW_SALARY_STAFF,
    "contractors": MonthLock.VIEW_SALARY_CONTRACTORS,
    "operators": MonthLock.VIEW_SALARY_OPERATORS,
    "fixed_payments": MonthLock.VIEW_SALARY_FIXED_PAYMENTS,
}


def _salary_decimal(request, field: str, emp_id) -> Decimal:
    raw = request.POST.get(f"{field}_{emp_id}", "").strip()
    if not raw:
        return Decimal(0)
    try:
        return Decimal(raw)
    except InvalidOperation:
        return Decimal(0)


def _row_missing_bank_details(row: dict) -> bool:
    """True if this employee can't go into the Bank Excel — either
    they're on Hold (nothing being paid out this month, so it's not
    actually missing anything) or their Employee record has no Account
    No / IFSC Code on file yet. Shared by _salary_context (to flag it on
    the page) and _write_salary_bank_sheet (to actually skip the row)."""
    if row["hold"]:
        return False
    emp = row["employee"]
    return not emp.account_no or not emp.ifsc_code


def _salary_context(current: date_cls) -> dict:
    """Computes one month's Salary page context — five tabs (Company
    Workers/Helpers/Staff/Contractors/Operators), each a bulk-editable
    table of that month's payroll adjustments (Adjust Days/Deductions/
    Additions, plus Manual Amount for Operators), with Gross/PF/ESI/NET
    computed via src/payroll.py from Employee.basic_salary/hra/da and
    this month's attendance. Contractors are prorated like Helpers/Staff
    but matched via Employee.department (the "Contractor" department)
    rather than subcategory. Shared by salary_view and
    salary_download_view so the .xlsx sheets can never drift from what
    the page shows."""
    year, month = current.year, current.month
    prev_date = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_date = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
    _, days_in_month = py_calendar.monthrange(year, month)
    month_start, month_end = current.replace(day=1), current.replace(day=days_in_month)

    working_days = payroll.working_days_in_month(year, month)
    special_day_counts = {
        r["day_type"]: r["n"]
        for r in SpecialDay.objects.filter(date__year=year, date__month=month)
        .values("day_type").annotate(n=Count("id"))
    }
    holiday_count = special_day_counts.get(SpecialDay.HOLIDAY, 0)
    paid_holiday_count = special_day_counts.get(SpecialDay.PAID_HOLIDAY, 0)
    comp_off_count = special_day_counts.get(SpecialDay.COMP_OFF, 0)

    adjustments = {a.employee_id: a for a in SalaryAdjustment.objects.filter(year=year, month=month)}
    paid_days_map = {
        r["employee_id"]: r["n"]
        for r in AttendanceRecord.objects.filter(
            date__year=year, date__month=month, status__in=["P", "PH", "CO"],
        ).values("employee_id").annotate(n=Count("id"))
    }

    def build_rows(employees, kind):
        rows = []
        for emp in employees:
            adj = adjustments.get(emp.id)
            paid_days = Decimal(paid_days_map.get(emp.id, 0))
            adjust_days = adj.adjust_days if adj else Decimal(0)
            deductions = adj.deductions if adj else Decimal(0)
            additions = adj.additions if adj else Decimal(0)
            manual_amount = adj.manual_amount if adj else None
            if kind == "company":
                calc = payroll.compute_company_worker_pay(
                    emp.basic_salary, emp.hra, emp.da, paid_days, working_days,
                    adjust_days, deductions, additions,
                    pf_enabled=emp.pf_enabled, esi_enabled=emp.esi_enabled,
                )
            elif kind == "operators":
                calc = payroll.compute_operator_pay(manual_amount, deductions, additions)
            elif kind == "fixed_payments":
                # A recurring flat amount set once on the Employee record
                # (Basic Salary — reused the same way Contractors reuse it
                # as a day rate), plus an optional manual top-up for a
                # specific month (e.g. a one-off extra charge) on top of
                # it, unlike Operators' manual_amount which stands in for
                # their whole pay.
                calc = payroll.compute_fixed_payment_pay(
                    emp.basic_salary, manual_amount, deductions, additions,
                )
            elif kind == "contractors":
                # Contractors are paid per day actually worked, not a fixed
                # monthly salary — Employee.basic_salary holds their day
                # rate for this group (reused rather than a separate field,
                # since it's otherwise unused for Contractors).
                calc = payroll.compute_daily_rate_pay(
                    emp.basic_salary, paid_days, adjust_days, deductions, additions,
                )
            else:
                calc = payroll.compute_prorated_pay(
                    emp.basic_salary, paid_days, working_days, adjust_days, deductions, additions,
                )
            hold = adj.hold if adj else False
            if hold:
                # Held pay is still computed in full above (so Gross/PF/
                # ESI/etc. keep reading correctly for records) — only NET
                # goes to 0, since nothing is actually being paid out this
                # month (see SalaryAdjustment.hold's help text).
                calc = {**calc, "net": Decimal(0)}
            rows.append({
                "employee": emp,
                "paid_days": paid_days,
                "adjust_days": adjust_days,
                "earned_days": round(paid_days + adjust_days, 2),
                "deductions": deductions,
                "additions": additions,
                "manual_amount": manual_amount,
                "hold": hold,
                "notes": adj.notes if adj else "",
                "calc": calc,
            })
        return rows

    def sum_rows(rows, kind):
        """Column totals for a tab's footer row — sums whatever numeric
        fields that tab's rows actually carry (employee Basic/HRA/DA for
        Company Workers, manual_amount for Operators, every payroll.py
        calc field for all of them), so each tab's template can reference
        totals.<field> the same way it references row.<field>."""
        totals = {
            "paid_days": sum((r["paid_days"] for r in rows), Decimal(0)),
            "adjust_days": sum((r["adjust_days"] for r in rows), Decimal(0)),
            "earned_days": sum((r["earned_days"] for r in rows), Decimal(0)),
            "deductions": sum((r["deductions"] for r in rows), Decimal(0)),
            "additions": sum((r["additions"] for r in rows), Decimal(0)),
        }
        if kind == "company":
            totals["basic_salary"] = sum((r["employee"].basic_salary for r in rows), Decimal(0))
            totals["da"] = sum((r["employee"].da for r in rows), Decimal(0))
            totals["hra"] = sum((r["employee"].hra for r in rows), Decimal(0))
        elif kind in ("helper", "staff", "contractors"):
            totals["basic_salary"] = sum((r["employee"].basic_salary for r in rows), Decimal(0))
        elif kind == "operators":
            totals["manual_amount"] = sum((r["manual_amount"] or Decimal(0) for r in rows), Decimal(0))
        elif kind == "fixed_payments":
            # Both figures apply here — the basic_salary-derived flat
            # payment AND a manual_amount on top of it for the odd
            # month it's more/less than usual (see build_rows/
            # compute_fixed_payment_pay).
            totals["basic_salary"] = sum((r["employee"].basic_salary for r in rows), Decimal(0))
            totals["manual_amount"] = sum((r["manual_amount"] or Decimal(0) for r in rows), Decimal(0))
        calc_keys = rows[0]["calc"].keys() if rows else []
        calc_totals = {k: Decimal(0) for k in calc_keys}
        for r in rows:
            for k in calc_keys:
                calc_totals[k] += r["calc"][k]
        totals["calc"] = {k: round(v, 2) for k, v in calc_totals.items()}
        return totals

    locked_views = set(
        MonthLock.objects.filter(year=year, month=month, view__in=_SALARY_LOCK_VIEWS.values())
        .values_list("view", flat=True)
    )
    lock_status = {tab_key: (view in locked_views) for tab_key, view in _SALARY_LOCK_VIEWS.items()}

    context = {
        "current": current,
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
        "working_days": working_days,
        "holiday_count": holiday_count,
        "paid_holiday_count": paid_holiday_count,
        "comp_off_count": comp_off_count,
        "lock_status": lock_status,
    }
    for tab_key, _label, subcategory in _SALARY_SUBCATEGORY_TABS:
        emp_filter = Q(subcategory__iexact=subcategory)
        # House Keeping employees don't consistently carry subcategory
        # "Helper" (one has it blank, per Employee data) but are paid the
        # same prorated way — folded into the Helpers tab by department
        # instead of leaving them off the Salary page (and Summary tab)
        # entirely.
        if tab_key == "helper":
            emp_filter |= Q(department__name__iexact="HOUSE KEEPING")
        employees = (
            Employee.objects.filter(emp_filter)
            .active_during(month_start, month_end).order_by("department__name", "name")
        )
        rows = build_rows(employees, tab_key)
        context[f"{tab_key}_rows"] = rows
        context[f"{tab_key}_totals"] = sum_rows(rows, tab_key)
    contractor_rows = build_rows(
        Employee.objects.filter(department__name__iexact="Contractor")
        .active_during(month_start, month_end).order_by("department__name", "name"),
        "contractors",
    )
    context["contractor_rows"] = contractor_rows
    context["contractor_totals"] = sum_rows(contractor_rows, "contractors")
    operator_rows = build_rows(
        Employee.objects.filter(category__iexact="Operator")
        .active_during(month_start, month_end).order_by("department__name", "name"),
        "operators",
    )
    context["operator_rows"] = operator_rows
    context["operator_totals"] = sum_rows(operator_rows, "operators")
    # Fixed Payments — recurring flat payments to specific individuals
    # (e.g. rent) that ride along on the Salary page for convenience but
    # aren't tied to attendance at all. The flat amount comes straight
    # from Employee.basic_salary (set once, not re-entered every month —
    # see build_rows's "fixed_payments" branch), unlike Operators' own
    # manual_amount which really is typed in fresh each month.
    fixed_payment_rows = build_rows(
        Employee.objects.filter(department__name__iexact="Fixed Payments")
        .active_during(month_start, month_end).order_by("department__name", "name"),
        "fixed_payments",
    )
    context["fixed_payment_rows"] = fixed_payment_rows
    context["fixed_payment_totals"] = sum_rows(fixed_payment_rows, "fixed_payments")

    # Summary tab — one row per salary group: headcount and NET total
    # (the one figure every group's rows carry in common — see
    # build_rows/sum_rows above — Gross/PF/ESI/Deductions/Additions vary
    # per group or aren't always meaningful across all five, so they're
    # left out of this cross-group rollup), plus a grand total row
    # summing every group together.
    summary_rows = []
    for label, rows_key, totals_key in [
        ("Company Workers", "company_rows", "company_totals"),
        ("Helpers", "helper_rows", "helper_totals"),
        ("Staff", "staff_rows", "staff_totals"),
        ("Contractors", "contractor_rows", "contractor_totals"),
        ("Operators", "operator_rows", "operator_totals"),
        ("Fixed Payments", "fixed_payment_rows", "fixed_payment_totals"),
    ]:
        totals = context[totals_key]
        summary_rows.append({
            "label": label,
            "count": len(context[rows_key]),
            "net": totals["calc"].get("net", Decimal(0)),
        })
    context["summary_rows"] = summary_rows
    context["summary_grand_total"] = {
        "count": sum(r["count"] for r in summary_rows),
        "net": sum((r["net"] for r in summary_rows), Decimal(0)),
    }

    context["missing_bank_details"] = [
        {"emp_code": r["employee"].code, "emp_name": r["employee"].name, "tab_label": tab_label}
        for tab_label, rows_key in [
            ("Company Workers", "company_rows"), ("Helpers", "helper_rows"), ("Staff", "staff_rows"),
            ("Contractors", "contractor_rows"), ("Operators", "operator_rows"),
            ("Fixed Payments", "fixed_payment_rows"),
        ]
        for r in context[rows_key]
        if _row_missing_bank_details(r)
    ]
    return context


@login_required
def salary_view(request):
    """Salary page — see _salary_context for the actual computation. Save
    is one upsert-by-(employee, year, month) per row, same bulk shape as
    mark_attendance_view's day-view save — one tab's table = one POST."""
    date_param = request.GET.get("date") or request.POST.get("date")
    current = _parse_month_date(date_param, date_cls.today().replace(day=1))
    year, month = current.year, current.month

    if request.method == "POST":
        tab = request.POST.get("tab", "")
        lock_view = _SALARY_LOCK_VIEWS.get(tab)
        if lock_view and MonthLock.objects.filter(year=year, month=month, view=lock_view).exists():
            _error(request, "This tab is locked for this month — unlock it first to make changes.")
            return redirect(f"{request.path}?date={current.isoformat()}#tab={tab}")

        employee_ids = request.POST.getlist("employee_id")
        saved = 0
        for emp_id in employee_ids:
            emp = Employee.objects.filter(id=emp_id).first()
            if not emp:
                continue
            manual_amount_raw = request.POST.get(f"manual_amount_{emp_id}", "").strip()
            manual_amount = None
            if manual_amount_raw:
                try:
                    manual_amount = Decimal(manual_amount_raw)
                except InvalidOperation:
                    manual_amount = None
            SalaryAdjustment.objects.update_or_create(
                employee=emp, year=year, month=month,
                defaults={
                    "adjust_days": _salary_decimal(request, "adjust_days", emp_id),
                    "deductions": _salary_decimal(request, "deductions", emp_id),
                    "additions": _salary_decimal(request, "additions", emp_id),
                    "manual_amount": manual_amount,
                    "hold": request.POST.get(f"hold_{emp_id}") == "on",
                    "notes": request.POST.get(f"notes_{emp_id}", "").strip(),
                },
            )
            saved += 1
        messages.success(request, f"Saved salary adjustments for {saved} employee(s) ({tab}).")
        logger.info(
            "Salary adjustments saved: tab=%s %s-%s saved=%s by user=%s",
            tab, year, month, saved, request.user,
        )
        return redirect(f"{request.path}?date={current.isoformat()}#tab={tab}")

    context = _salary_context(current)
    return render(request, "attendance/salary.html", context)


# Same background colors as the page's column-group shading (see
# table.simple th/td.grp-* in salary.html) — keyed by the same group names
# so a sheet's header/data cells can be filled to match the on-screen tab
# exactly. The Total row is deliberately NOT shaded per-group (the page's
# tfoot cells don't carry any grp-* class either — the whole row just gets
# a uniform pale yellow, same as every other download on this site).
_SALARY_GROUP_FILLS = {
    "fixed": "ECEEF0",
    "attendance": "B4C7E7",
    "deduction-emp": "FFF2CC",
    "deduction-employer": "FFE0B2",
    "highlight": "FFFF00",
}


# (tab key, sheet title, headers, row-builder, totals-row-builder) for each
# of the five Salary tabs' .xlsx sheets — see salary_download_view. Kept as
# plain functions (not a shared column set) since Company Workers' PF/ESI
# breakdown, Staff's TDS? column and Operators' Manual Amount/Notes each
# need their own shape; every row/total list must line up with its own
# headers list position-for-position, and so must col_groups (one entry
# per header — a key into _SALARY_GROUP_FILLS, or None for an unshaded
# column) — see _write_salary_sheet.
def _salary_company_sheet_rows(rows, totals):
    headers = [
        "Code", "Employee", "Basic", "DA", "Basic+DA", "HRA", "Gross",
        "Paid Days", "Adjust Days", "Earned Days",
        "Earned Basic+DA", "Earned HRA", "Earned Total Wages",
        "PF (Employee)", "ESI (Employee)", "PF+ESI (Employee)",
        "PF (Employer)", "ESI (Employer)", "PF+ESI (Employer)",
        "Deductions", "Total Deductions", "Additions", "Hold", "NET",
    ]
    col_groups = [
        None, None, "fixed", "fixed", "fixed", "fixed", "fixed",
        "attendance", "attendance", "attendance",
        None, None, None,
        "deduction-emp", "deduction-emp", "deduction-emp",
        "deduction-employer", "deduction-employer", "deduction-employer",
        "deduction-emp", "deduction-emp", None, None, "highlight",
    ]
    data_rows = [
        [
            r["employee"].code, r["employee"].name,
            float(r["employee"].basic_salary), float(r["employee"].da), float(r["calc"]["basic_da"]),
            float(r["employee"].hra), float(r["calc"]["gross"]),
            float(r["paid_days"]), float(r["adjust_days"]), float(r["calc"]["earned_days"]),
            float(r["calc"]["earned_basic_da"]), float(r["calc"]["earned_hra"]), float(r["calc"]["earned_total"]),
            float(r["calc"]["pf"]), float(r["calc"]["esi"]), float(r["calc"]["pf_esi_employee"]),
            float(r["calc"]["pf_employer"]), float(r["calc"]["esi_employer"]), float(r["calc"]["pf_esi_employer"]),
            float(r["deductions"]), float(r["calc"]["total_deduction"]), float(r["additions"]),
            "Yes" if r["hold"] else "No", float(r["calc"]["net"]),
        ]
        for r in rows
    ]
    # totals["calc"] is {} when the tab has zero rows this month (see
    # sum_rows) — .get(..., 0) so an empty tab still downloads a Total row
    # of zeros instead of a KeyError, matching the template's silent
    # {{ totals.calc.net }} on a missing key.
    c = totals["calc"]
    total_row = [
        "", "Total",
        float(totals["basic_salary"]), float(totals["da"]), float(c.get("basic_da", 0)),
        float(totals["hra"]), float(c.get("gross", 0)),
        float(totals["paid_days"]), float(totals["adjust_days"]), float(c.get("earned_days", 0)),
        float(c.get("earned_basic_da", 0)), float(c.get("earned_hra", 0)), float(c.get("earned_total", 0)),
        float(c.get("pf", 0)), float(c.get("esi", 0)), float(c.get("pf_esi_employee", 0)),
        float(c.get("pf_employer", 0)), float(c.get("esi_employer", 0)), float(c.get("pf_esi_employer", 0)),
        float(totals["deductions"]), float(c.get("total_deduction", 0)), float(totals["additions"]),
        "", float(c.get("net", 0)),
    ]
    return headers, col_groups, data_rows, total_row


def _salary_prorated_sheet_rows(rows, totals, with_tds: bool, basic_salary_label: str = "Basic Salary"):
    """Helpers/Contractors (with_tds=False) and Staff (with_tds=True) —
    same shape (a per-employee rate x Earned Days), Staff just adds a
    TDS? column. basic_salary_label lets Contractors' sheet say "Day
    Rate" instead, since Employee.basic_salary means something different
    there (see build_rows/compute_daily_rate_pay in _salary_context)."""
    headers = [
        "Code", "Employee", basic_salary_label, "Paid Days", "Adjust Days", "Earned Days",
        "Earned Salary", "Deductions", "Additions", "Hold",
    ]
    col_groups = [None, None, "fixed", "attendance", "attendance", "attendance", None, "deduction-emp", None, None]
    if with_tds:
        headers.append("TDS?")
        col_groups.append(None)
    headers.append("NET")
    col_groups.append("highlight")

    data_rows = []
    for r in rows:
        row = [
            r["employee"].code, r["employee"].name, float(r["employee"].basic_salary),
            float(r["paid_days"]), float(r["adjust_days"]), float(r["earned_days"]),
            float(r["calc"]["earned_salary"]), float(r["deductions"]), float(r["additions"]),
            "Yes" if r["hold"] else "No",
        ]
        if with_tds:
            row.append("Yes" if r["employee"].tds_enabled else "No")
        row.append(float(r["calc"]["net"]))
        data_rows.append(row)

    # totals["calc"] is {} when the tab has zero rows this month — see the
    # matching comment in _salary_company_sheet_rows.
    c = totals["calc"]
    total_row = [
        "", "Total", float(totals["basic_salary"]),
        float(totals["paid_days"]), float(totals["adjust_days"]), float(totals["earned_days"]),
        float(c.get("earned_salary", 0)), float(totals["deductions"]), float(totals["additions"]), "",
    ]
    if with_tds:
        total_row.append("")
    total_row.append(float(c.get("net", 0)))
    return headers, col_groups, data_rows, total_row


def _salary_summary_sheet_rows(summary_rows, grand_total):
    headers = ["Salary Group", "No. of People", "NET"]
    col_groups = [None, None, "highlight"]
    data_rows = [[r["label"], r["count"], float(r["net"])] for r in summary_rows]
    total_row = ["Total", grand_total["count"], float(grand_total["net"])]
    return headers, col_groups, data_rows, total_row


def _salary_operator_sheet_rows(rows, totals):
    headers = ["Code", "Employee", "Earned Days", "Manual Amount", "Deductions", "Additions", "Hold", "Notes", "NET"]
    col_groups = [None, None, "attendance", "fixed", "deduction-emp", None, None, None, "highlight"]
    data_rows = [
        [
            r["employee"].code, r["employee"].name, float(r["earned_days"]),
            float(r["manual_amount"] or 0), float(r["deductions"]), float(r["additions"]),
            "Yes" if r["hold"] else "No", r["notes"], float(r["calc"]["net"]),
        ]
        for r in rows
    ]
    # totals["calc"] is {} when the tab has zero rows this month — see the
    # matching comment in _salary_company_sheet_rows.
    total_row = [
        "", "Total", float(totals["earned_days"]), float(totals["manual_amount"]),
        float(totals["deductions"]), float(totals["additions"]), "", "", float(totals["calc"].get("net", 0)),
    ]
    return headers, col_groups, data_rows, total_row


def _salary_fixed_payment_sheet_rows(rows, totals):
    """Fixed Payments: a recurring flat amount from Employee.basic_salary
    (labeled "Fixed Payment") plus an optional Manual Amount top-up for a
    specific month, both feeding NET — see compute_fixed_payment_pay.
    "Code"/"Employee" are relabeled "Purpose"/"Name" since these payees
    (rent, etc.) aren't really employees with a code."""
    headers = ["Purpose", "Name", "Fixed Payment", "Manual Amount", "Deductions", "Additions", "Hold", "Notes", "NET"]
    col_groups = [None, None, "fixed", "fixed", "deduction-emp", None, None, None, "highlight"]
    data_rows = [
        [
            r["employee"].code, r["employee"].name, float(r["employee"].basic_salary),
            float(r["manual_amount"] or 0), float(r["deductions"]), float(r["additions"]),
            "Yes" if r["hold"] else "No", r["notes"], float(r["calc"]["net"]),
        ]
        for r in rows
    ]
    # totals["calc"] is {} when the tab has zero rows this month — see the
    # matching comment in _salary_company_sheet_rows.
    total_row = [
        "", "Total", float(totals["basic_salary"]), float(totals["manual_amount"]),
        float(totals["deductions"]), float(totals["additions"]), "", "", float(totals["calc"].get("net", 0)),
    ]
    return headers, col_groups, data_rows, total_row


def _write_salary_sheet(
    ws, label: str, month_label: str, headers: list, col_groups: list, data_rows: list, total_row: list,
) -> None:
    """Fills in one Salary tab's sheet — title row, header row, data rows,
    then a bold yellow Total row (matching the OT Details/Cash Withdrawal
    downloads' styling — see _apply_grid_borders). Header and data cells
    are shaded per col_groups (one entry per header, a key into
    _SALARY_GROUP_FILLS or None) so the sheet reads the same color-coded
    columns as the on-screen tab; an unshaded header still gets a plain
    grey fill so it reads as a header, but unshaded data cells are left
    white, same as the page. The Total row is deliberately uniform (no
    per-group shading), matching the page's tfoot."""
    from openpyxl.styles import Font, PatternFill
    import openpyxl

    ws.append([f"{label} — {month_label}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(headers)
    header_row_num = ws.max_row
    default_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    group_fills = {
        key: PatternFill(start_color=color, end_color=color, fill_type="solid")
        for key, color in _SALARY_GROUP_FILLS.items()
    }
    for cell, group in zip(ws[header_row_num], col_groups):
        cell.font = Font(bold=True)
        cell.fill = group_fills.get(group, default_header_fill)

    for row in data_rows:
        ws.append(row)
        for cell, group in zip(ws[ws.max_row], col_groups):
            if group:
                cell.fill = group_fills[group]
                if group == "highlight":
                    cell.font = Font(bold=True)

    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF9B0", end_color="FFF9B0", fill_type="solid")

    _apply_grid_borders(ws, header_row_num, ws.max_row, len(headers))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14


@login_required
def salary_download_view(request):
    """Downloads one month's Salary page as a single .xlsx with one sheet
    per selected tab (see _SALARY_TAB_KEYS; ?tabs=company&tabs=helper&...
    from the page's tab-picker, or every tab if none were given) — via
    the same _salary_context used by the page, so the sheets and the
    on-screen tabs can never drift apart."""
    import openpyxl

    date_param = request.GET.get("date")
    current = _parse_month_date(date_param, date_cls.today().replace(day=1))
    context = _salary_context(current)
    month_label = f"{context['month_name']} {context['year']}"

    sheet_builders = {
        "summary": lambda: _salary_summary_sheet_rows(context["summary_rows"], context["summary_grand_total"]),
        "company": lambda: _salary_company_sheet_rows(context["company_rows"], context["company_totals"]),
        "helper": lambda: _salary_prorated_sheet_rows(context["helper_rows"], context["helper_totals"], with_tds=False),
        "staff": lambda: _salary_prorated_sheet_rows(context["staff_rows"], context["staff_totals"], with_tds=True),
        "contractors": lambda: (
            _salary_prorated_sheet_rows(
                context["contractor_rows"], context["contractor_totals"], with_tds=False,
                basic_salary_label="Day Rate",
            )
        ),
        "operators": lambda: _salary_operator_sheet_rows(context["operator_rows"], context["operator_totals"]),
        "fixed_payments": lambda: (
            _salary_fixed_payment_sheet_rows(context["fixed_payment_rows"], context["fixed_payment_totals"])
        ),
    }
    # "summary" isn't one of _SALARY_TAB_KEYS (that list is shared with the
    # Bank download, which has no use for an aggregate sheet) — handled
    # here instead, and placed first to match its position on the page
    # and in the Telegram/Email pickers.
    selected_keys = set(request.GET.getlist("tabs")) or ({key for key, _, _ in _SALARY_TAB_KEYS} | {"summary"})
    selected = []
    if "summary" in selected_keys:
        selected.append(("Summary", "summary"))
    selected += [(label, key) for key, label, _ in _SALARY_TAB_KEYS if key in selected_keys]
    if not selected:
        _error(request, "Select at least one tab to download.")
        return redirect(f"{reverse('salary')}?date={current.isoformat()}")

    wb = openpyxl.Workbook()
    first = True
    for label, key in selected:
        ws = wb.active if first else wb.create_sheet(label)
        if first:
            ws.title = label
            first = False
        headers, col_groups, data_rows, total_row = sheet_builders[key]()
        _write_salary_sheet(ws, label, month_label, headers, col_groups, data_rows, total_row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"salary-{context['year']}-{context['month']:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# Company name printed into every bank sheet's REMITTER'S NAME column
# (see _salary_bank_row) — the account number next to it is deliberately
# left blank in every row, same as the July template; finance fills that
# in centrally when uploading to the bank's own portal.
_SALARY_BANK_REMITTER_NAME = "MV INDUSTRIAL CORPORATION"

# Column layout lifted directly from the July template's per-department
# Bank Sheet tabs (Workers_Bank_Sheet, Helpers_Bank_Sheet, Op_Bank_Sheet,
# etc.) — a blank spacer right after EMP NAME, then a literal "~" spacer
# between every other column pair, and two columns (10, 2) that are
# fixed constants in every row, including the header. Replicated exactly,
# typos in "BENFICIARY'S" included, since this file is meant to be
# uploaded as-is to the bank's NEFT bulk-transfer portal, which expects
# that exact shape.
_SALARY_BANK_SHEET_HEADER = [
    "EMP NAME", None, "TRANSFER TYPE", "~", "REMITTER'S ACCOUNT", "~", "REMITTER'S NAME", "~",
    "IFSC CODE", "~", "ACCOUNT NO", "~", 10, "~", "AMOUNT OF TRANSFER", "~",
    "BENEFICIARY'S NAME", "~", "BENFICIARY'S BRANCH NAME", "~", 2, "~", "BENFICIARY'S BANK NAME",
]


def _salary_bank_row(emp: Employee, amount) -> list:
    """One data row for a Bank Sheet — see _SALARY_BANK_SHEET_HEADER for
    the column layout this must line up with position-for-position."""
    return [
        emp.name, None, "NEFT", "~", None, "~", _SALARY_BANK_REMITTER_NAME, "~",
        emp.ifsc_code, "~", emp.account_no, "~", 10, "~", float(amount), "~",
        emp.account_name or emp.name, "~", emp.branch, "~", 2, "~", emp.bank_name,
    ]


def _write_salary_bank_sheet(ws, rows: list) -> None:
    """Fills in one tab's Bank Sheet — header starts at row 1 with no
    title above it (unlike this app's other downloads), matching the
    July template's bank sheets exactly since this file gets uploaded
    straight to the bank, not read by a person. Skips employees on Hold
    and employees missing Account No/IFSC Code (see
    _row_missing_bank_details — those are flagged on the Salary page
    itself instead, via _salary_context's missing_bank_details)."""
    from openpyxl.styles import Font

    ws.append(_SALARY_BANK_SHEET_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        if _row_missing_bank_details(r) or r["hold"]:
            continue
        ws.append(_salary_bank_row(r["employee"], r["calc"]["net"]))

    ws.column_dimensions["A"].width = 22
    for col in ("C", "E", "G", "I", "K", "M", "O", "Q", "S", "U", "W"):
        ws.column_dimensions[col].width = 16
    for col in ("B", "D", "F", "H", "J", "L", "N", "P", "R", "T", "V"):
        ws.column_dimensions[col].width = 3


@login_required
def salary_bank_download_view(request):
    """Downloads one month's Salary page as a NEFT bulk-transfer .xlsx —
    one sheet per selected tab (see _SALARY_TAB_KEYS; ?tabs=company&
    tabs=helper&... from the page's tab-picker, or every tab if none were
    given), same column layout as the July template's Bank Sheet tabs
    (see _write_salary_bank_sheet). Employees on Hold are left out of the
    file (nothing being paid out this month) — but if anyone else in a
    selected tab is missing bank details, the whole download is refused
    instead of silently going out short one person's transfer; fix the
    missing Employee record(s) (also listed on the Salary page itself,
    see missing_bank_details in _salary_context) and try again."""
    import openpyxl

    date_param = request.GET.get("date")
    current = _parse_month_date(date_param, date_cls.today().replace(day=1))
    context = _salary_context(current)

    selected_keys = set(request.GET.getlist("tabs")) or {key for key, _, _ in _SALARY_TAB_KEYS}
    selected = [(label, rows_key) for key, label, rows_key in _SALARY_TAB_KEYS if key in selected_keys]
    if not selected:
        _error(request, "Select at least one tab to download the Bank Excel for.")
        return redirect(f"{reverse('salary')}?date={current.isoformat()}")

    selected_labels = {label for label, _ in selected}
    missing = [m for m in context["missing_bank_details"] if m["tab_label"] in selected_labels]
    if missing:
        names = ", ".join(f"{m['emp_code']} {m['emp_name']} ({m['tab_label']})" for m in missing)
        _error(
            request,
            f"Bank Excel not downloaded — missing bank details for: {names}. "
            "Add their Account No/IFSC Code first.",
        )
        return redirect(f"{reverse('salary')}?date={current.isoformat()}")

    wb = openpyxl.Workbook()
    first = True
    for label, rows_key in selected:
        ws = wb.active if first else wb.create_sheet(label)
        if first:
            ws.title = label
            first = False
        _write_salary_bank_sheet(ws, context[rows_key])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"salary-bank-{context['year']}-{context['month']:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def cash_withdrawal_view(request):
    """Monthly cash-register log — money withdrawn for fixed purposes
    outside of payroll (petty cash, office expenses, an advance, etc.), each a
    simple Purpose/Description/Amount row logged against a month as a
    whole (no specific day), listed and totaled like Salary/OT Details.
    Not employee-linked — this is general cash bookkeeping, not a payroll
    adjustment."""
    date_param = request.GET.get("date") or request.POST.get("date")
    current = _parse_month_date(date_param, date_cls.today().replace(day=1))
    year, month = current.year, current.month
    prev_date = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_date = (current.replace(day=28) + timedelta(days=4)).replace(day=1)

    if request.method == "POST":
        action = request.POST.get("action", "add")
        if action == "delete":
            CashWithdrawal.objects.filter(id=request.POST.get("entry_id")).delete()
            messages.success(request, "Entry deleted.")
            return redirect(f"{request.path}?date={current.isoformat()}")

        purpose = request.POST.get("purpose", "").strip()
        amount_raw = request.POST.get("amount", "").strip()
        description = request.POST.get("description", "").strip()
        try:
            amount = Decimal(amount_raw)
        except InvalidOperation:
            _error(request, "Enter a valid amount.")
            return redirect(f"{request.path}?date={current.isoformat()}")
        if not purpose:
            _error(request, "Purpose is required.")
            return redirect(f"{request.path}?date={current.isoformat()}")

        if action == "edit":
            entry = CashWithdrawal.objects.filter(id=request.POST.get("entry_id")).first()
            if not entry:
                _error(request, "Entry not found — it may have already been deleted.")
                return redirect(f"{request.path}?date={current.isoformat()}")
            entry.purpose = purpose
            entry.amount = amount
            entry.description = description
            entry.save()
            messages.success(request, "Cash withdrawal updated.")
            logger.info(
                "Cash withdrawal updated: id=%s %s-%s %s %s by user=%s",
                entry.id, year, month, purpose, amount, request.user,
            )
        else:
            CashWithdrawal.objects.create(year=year, month=month, purpose=purpose, amount=amount, description=description)
            messages.success(request, "Cash withdrawal recorded.")
            logger.info("Cash withdrawal recorded: %s-%s %s %s by user=%s", year, month, purpose, amount, request.user)
        return redirect(f"{request.path}?date={current.isoformat()}")

    entries = CashWithdrawal.objects.filter(year=year, month=month)
    total_amount = sum((e.amount for e in entries), Decimal(0))

    context = {
        "current": current,
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
        "entries": entries,
        "total_amount": total_amount,
    }
    return render(request, "attendance/cash_withdrawal.html", context)


@login_required
def cash_withdrawal_download_view(request):
    """Downloads one month's Cash Withdrawal log (Purpose/Description/
    Amount, same rows as the page) as a single-sheet .xlsx with a bold
    yellow Total row, matching the OT Details downloads' styling."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    current = _parse_month_date(request.GET.get("date"), date_cls.today().replace(day=1))
    year, month = current.year, current.month
    entries = CashWithdrawal.objects.filter(year=year, month=month)
    total_amount = sum((e.amount for e in entries), Decimal(0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Withdrawal"

    month_label = f"{py_calendar.month_name[month]} {year}"
    ws.append([f"Cash Withdrawal — {month_label}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    headers = ["Purpose", "Description", "Amount"]
    ws.append(headers)
    header_row_num = ws.max_row
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for entry in entries:
        ws.append([entry.purpose, entry.description, float(entry.amount)])

    ws.append(["Total", "", float(total_amount)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF9B0", end_color="FFF9B0", fill_type="solid")

    _apply_grid_borders(ws, header_row_num, ws.max_row, len(headers))

    widths = [24, 36, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"cash-withdrawal-{year}-{month:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _cash_register_signed(entry: CashRegisterEntry) -> Decimal:
    return entry.amount if entry.entry_type == CashRegisterEntry.TYPE_IN else -entry.amount


def _cash_register_context(current: date_cls) -> dict:
    """Computes one month's Cash Register — a day-by-day petty-cash
    ledger (Cash In/Cash Out) with a running balance per row. A month's
    opening balance is just the signed total of every entry dated before
    that month started, so the register carries over month to month with
    nothing stored beyond the entries themselves — see
    CashRegisterEntry's own docstring for why. Shared by
    cash_register_view and cash_register_download_view so the download
    can never drift from what the page shows."""
    year, month = current.year, current.month
    prev_date = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_date = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_start = current.replace(day=1)
    _, days_in_month = py_calendar.monthrange(year, month)
    month_end = current.replace(day=days_in_month)

    opening_balance = sum(
        (_cash_register_signed(e) for e in CashRegisterEntry.objects.filter(date__lt=month_start)),
        Decimal(0),
    )

    running = opening_balance
    total_in = Decimal(0)
    total_out = Decimal(0)
    rows = []
    for entry in CashRegisterEntry.objects.filter(date__gte=month_start, date__lte=month_end):
        running += _cash_register_signed(entry)
        if entry.entry_type == CashRegisterEntry.TYPE_IN:
            total_in += entry.amount
        else:
            total_out += entry.amount
        rows.append({"entry": entry, "balance": running})

    return {
        "current": current,
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
        "rows": rows,
        "opening_balance": opening_balance,
        "closing_balance": running,
        "total_in": total_in,
        "total_out": total_out,
    }


@login_required
def cash_register_view(request):
    """Petty cash register — a day-by-day ledger of Cash In (e.g. money
    withdrawn from the bank into the office cash box) and Cash Out
    (an expense/payment made from it), with a running balance that
    carries over month to month (see _cash_register_context). Kept as
    its own page next to the simpler, undated Cash Withdrawal log under
    the same "Cash" nav menu — the two serve different habits, not one
    replacing the other."""
    date_param = request.GET.get("date") or request.POST.get("date")
    current = _parse_month_date(date_param, date_cls.today().replace(day=1))

    if request.method == "POST":
        action = request.POST.get("action", "add")
        if action == "delete":
            CashRegisterEntry.objects.filter(id=request.POST.get("entry_id")).delete()
            messages.success(request, "Entry deleted.")
            return redirect(f"{request.path}?date={current.isoformat()}")

        entry_type = request.POST.get("entry_type", "")
        purpose = request.POST.get("purpose", "").strip()
        amount_raw = request.POST.get("amount", "").strip()
        description = request.POST.get("description", "").strip()
        date_raw = request.POST.get("entry_date", "").strip()

        if entry_type not in (CashRegisterEntry.TYPE_IN, CashRegisterEntry.TYPE_OUT):
            _error(request, "Select Cash In or Cash Out.")
            return redirect(f"{request.path}?date={current.isoformat()}")
        if not purpose:
            _error(request, "Purpose is required.")
            return redirect(f"{request.path}?date={current.isoformat()}")
        try:
            amount = Decimal(amount_raw)
        except InvalidOperation:
            _error(request, "Enter a valid amount.")
            return redirect(f"{request.path}?date={current.isoformat()}")
        try:
            entry_date = date_cls.fromisoformat(date_raw) if date_raw else date_cls.today()
        except ValueError:
            _error(request, "Enter a valid date.")
            return redirect(f"{request.path}?date={current.isoformat()}")

        if action == "edit":
            entry = CashRegisterEntry.objects.filter(id=request.POST.get("entry_id")).first()
            if not entry:
                _error(request, "Entry not found — it may have already been deleted.")
                return redirect(f"{request.path}?date={current.isoformat()}")
            entry.date = entry_date
            entry.entry_type = entry_type
            entry.purpose = purpose
            entry.description = description
            entry.amount = amount
            entry.save()
            messages.success(request, "Cash register entry updated.")
            logger.info(
                "Cash register entry updated: id=%s %s %s %s %s by user=%s",
                entry.id, entry_date, entry_type, purpose, amount, request.user,
            )
        else:
            CashRegisterEntry.objects.create(
                date=entry_date, entry_type=entry_type, purpose=purpose, amount=amount, description=description,
            )
            messages.success(request, "Cash register entry recorded.")
            logger.info(
                "Cash register entry recorded: %s %s %s %s by user=%s",
                entry_date, entry_type, purpose, amount, request.user,
            )
        # Land on whichever month the entry actually belongs to, not
        # necessarily the month that was being viewed (a backdated/
        # postdated entry, or one edited into a different month, lands
        # elsewhere on purpose).
        return redirect(f"{request.path}?date={entry_date.replace(day=1).isoformat()}")

    context = _cash_register_context(current)
    return render(request, "attendance/cash_register.html", context)


@login_required
def cash_register_download_view(request):
    """Downloads one month's Cash Register as a single-sheet .xlsx —
    Date/Type/Purpose/Description/Cash In/Cash Out/Balance, with the
    month's opening and closing balance rows — via the same
    _cash_register_context used by the page."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    current = _parse_month_date(request.GET.get("date"), date_cls.today().replace(day=1))
    context = _cash_register_context(current)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Register"

    month_label = f"{context['month_name']} {context['year']}"
    ws.append([f"Cash Register — {month_label}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    headers = ["Date", "Type", "Purpose", "Description", "Cash In", "Cash Out", "Balance"]
    ws.append(headers)
    header_row_num = ws.max_row
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.append(["", "", "Opening Balance", "", "", "", float(context["opening_balance"])])
    for cell in ws[ws.max_row]:
        cell.font = Font(italic=True)

    for row in context["rows"]:
        entry = row["entry"]
        is_in = entry.entry_type == CashRegisterEntry.TYPE_IN
        ws.append([
            entry.date.strftime("%Y-%m-%d"), entry.get_entry_type_display(), entry.purpose, entry.description,
            float(entry.amount) if is_in else None,
            float(entry.amount) if not is_in else None,
            float(row["balance"]),
        ])

    ws.append([
        "", "", "Closing Balance", "",
        float(context["total_in"]), float(context["total_out"]), float(context["closing_balance"]),
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF9B0", end_color="FFF9B0", fill_type="solid")

    _apply_grid_borders(ws, header_row_num, ws.max_row, len(headers))

    widths = [12, 10, 22, 30, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"cash-register-{context['year']}-{context['month']:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _leave_ledger_context(current: date_cls) -> dict:
    """Computes one month's EL/Comp-Off ledger — one row per active Staff
    employee (see LeaveLedgerEntry for the accrual rules). If HR has
    already posted this month for an employee, their saved figures are
    shown as-is; otherwise this computes a live preview of what posting
    would produce, so the page always shows real numbers whether or not
    the month has been closed yet. Shared by leave_ledger_view's GET and
    its "Post this month" bulk action, so the preview and what actually
    gets saved can never disagree."""
    year, month = current.year, current.month
    prev_date = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_date = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_start = current.replace(day=1)
    _, days_in_month = py_calendar.monthrange(year, month)
    month_end = current.replace(day=days_in_month)

    staff = list(
        Employee.objects.filter(subcategory__iexact="Staff")
        .active_during(month_start, month_end).order_by("name")
    )

    # Most recent LeaveLedgerEntry strictly before this month, per
    # employee — ascending order plus a plain dict overwrite means the
    # last one written per employee is the latest one before (year,
    # month), without a separate "max" query per employee.
    prior_by_emp = {}
    for entry in (
        LeaveLedgerEntry.objects.filter(Q(year__lt=year) | Q(year=year, month__lt=month))
        .order_by("year", "month")
    ):
        prior_by_emp[entry.employee_id] = entry

    this_month_by_emp = {e.employee_id: e for e in LeaveLedgerEntry.objects.filter(year=year, month=month)}

    # Spent EL — an actual day taken off — is just an AttendanceRecord
    # marked status="EL" (see AttendanceRecord.STATUS_CHOICES), same as
    # any other leave type; counted straight from the DB rather than the
    # pandas pipeline since it doesn't depend on SpecialDay/OT at all.
    el_taken_map: dict = {}
    el_taken_dates_map: dict = {}
    for r in (
        AttendanceRecord.objects.filter(date__year=year, date__month=month, status="EL")
        .order_by("date").values("employee_id", "date")
    ):
        el_taken_map[r["employee_id"]] = el_taken_map.get(r["employee_id"], 0) + 1
        el_taken_dates_map.setdefault(r["employee_id"], []).append(r["date"])

    full_ot_map: dict = {}
    comp_off_map: dict = {}
    daily_all = _load_daily_data()
    if not daily_all.empty:
        daily = daily_all[(daily_all["date"].dt.year == year) & (daily_all["date"].dt.month == month)]
        if not daily.empty and "subcategory" in daily.columns:
            special_days, downgraded_special_days, skip_special_days = _special_days_and_downgrades()
            daily = metrics.apply_special_days(daily, special_days, downgraded_special_days, skip_special_days)
            staff_daily = daily[daily["subcategory"] == "Staff"]
            if not staff_daily.empty:
                # el_day_credit (1.0 full day / 0.5 half day, by hours
                # worked) comes from overtime_view — the same computation
                # the OT page's EL Days column uses (see
                # _apply_staff_ot_display) — rather than a plain count of
                # special_worked days, so the two pages can never disagree
                # about how many EL days a given month's holidays earned.
                shift_ot_table = metrics.overtime_view(daily)
                if not shift_ot_table.empty:
                    staff_codes = set(staff_daily["emp_code"])
                    full_ot_rows = shift_ot_table[
                        shift_ot_table["emp_code"].isin(staff_codes) & (shift_ot_table["full_day_ot"] == 1)
                    ]
                    full_ot_map = full_ot_rows.groupby("emp_code")["el_day_credit"].sum().to_dict()
                excess = (staff_daily["work_hours"] - float(LeaveLedgerEntry.COMP_OFF_HOUR_THRESHOLD)).clip(lower=0)
                comp_off_map = staff_daily.assign(_excess=excess).groupby("emp_code")["_excess"].sum().to_dict()

    rows = []
    for emp in staff:
        posted = this_month_by_emp.get(emp.id)
        prior = prior_by_emp.get(emp.id)
        prior_el = prior.el_balance_after if prior else Decimal(0)
        prior_comp_off = prior.comp_off_balance_after if prior else Decimal(0)

        # Always derived live from AttendanceRecord, even for an
        # already-posted month — el_taken itself is frozen in the ledger
        # row like everything else, but the specific dates are just a
        # display aid, and showing a mismatch (if attendance was edited
        # after posting) is more useful than hiding it.
        el_taken_dates = el_taken_dates_map.get(emp.id, [])

        if posted:
            rows.append({
                "employee": emp, "posted": True, "entry_id": posted.id,
                "prior_el_balance": prior_el, "prior_comp_off_balance": prior_comp_off,
                "full_ot_days": posted.full_ot_days,
                "el_credited": posted.el_credited, "el_encashed": posted.el_encashed,
                "el_taken": posted.el_taken,
                "el_taken_dates": el_taken_dates,
                "el_balance_after": posted.el_balance_after,
                "comp_off_hours_earned": posted.comp_off_hours_earned,
                "comp_off_balance_after": posted.comp_off_balance_after,
                "is_manual": posted.is_manual,
                "notes": posted.notes,
            })
            continue

        full_ot_days = Decimal(str(round(float(full_ot_map.get(emp.code, 0)), 2)))
        comp_off_hours_earned = Decimal(str(round(float(comp_off_map.get(emp.code, 0)), 2)))
        el_taken = Decimal(el_taken_map.get(emp.id, 0))

        room = LeaveLedgerEntry.EL_CAP - prior_el
        if room < 0:
            room = Decimal(0)
        el_credited = min(full_ot_days, room)
        el_encashed = full_ot_days - el_credited

        rows.append({
            "employee": emp, "posted": False, "entry_id": None,
            "prior_el_balance": prior_el, "prior_comp_off_balance": prior_comp_off,
            "full_ot_days": full_ot_days,
            "el_credited": el_credited, "el_encashed": el_encashed,
            "el_taken": el_taken,
            "el_taken_dates": el_taken_dates,
            "el_balance_after": prior_el + el_credited - el_taken,
            "comp_off_hours_earned": comp_off_hours_earned,
            "comp_off_balance_after": prior_comp_off + comp_off_hours_earned,
            "is_manual": False,
            "notes": "",
        })

    return {
        "current": current,
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
        "rows": rows,
        "el_cap": LeaveLedgerEntry.EL_CAP,
        "comp_off_threshold": LeaveLedgerEntry.COMP_OFF_HOUR_THRESHOLD,
    }


@login_required
def leave_ledger_view(request):
    """EL (Earned Leave) / Comp-Off page for Staff — see
    _leave_ledger_context for the accrual computation and
    LeaveLedgerEntry for the rules. "Post this month" (bulk, for every
    Staff employee at once) saves the current preview as that month's
    permanent figures; re-posting recomputes and overwrites only that
    month's row, so if attendance for an already-posted month changes
    later, re-post it — and any later month already posted will need
    re-posting too, in order, since each month's balance carries from
    the one before it. "Edit" on one employee's row instead sets that
    month's balances by hand (is_manual=True) — the way to seed a
    starting EL/Comp-Off balance from before this ledger existed, or to
    hand-correct a mistake."""
    date_param = request.GET.get("date") or request.POST.get("date")
    current = _parse_month_date(date_param, date_cls.today().replace(day=1))
    year, month = current.year, current.month

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "delete":
            LeaveLedgerEntry.objects.filter(id=request.POST.get("entry_id")).delete()
            messages.success(request, "Entry removed — this month will show a fresh preview again.")
            return redirect(f"{request.path}?date={current.isoformat()}")

        if action == "edit":
            emp = Employee.objects.filter(id=request.POST.get("employee_id")).first()
            if not emp:
                _error(request, "Employee not found.")
                return redirect(f"{request.path}?date={current.isoformat()}")
            try:
                el_balance_after = Decimal(request.POST.get("el_balance_after", "").strip())
                comp_off_balance_after = Decimal(request.POST.get("comp_off_balance_after", "").strip())
            except InvalidOperation:
                _error(request, "Enter valid numbers for the EL and Comp-Off balances.")
                return redirect(f"{request.path}?date={current.isoformat()}")
            notes = request.POST.get("notes", "").strip()

            LeaveLedgerEntry.objects.update_or_create(
                employee=emp, year=year, month=month,
                defaults={
                    "full_ot_days": 0, "el_credited": 0, "el_encashed": 0, "el_taken": 0,
                    "el_balance_after": el_balance_after,
                    "comp_off_hours_earned": 0, "comp_off_balance_after": comp_off_balance_after,
                    "is_manual": True, "notes": notes,
                },
            )
            messages.success(request, f"Manually set {emp.name}'s balances for {py_calendar.month_name[month]} {year}.")
            logger.info(
                "Leave ledger manual entry: emp=%s %s-%s EL=%s CompOff=%sh by user=%s",
                emp.code, year, month, el_balance_after, comp_off_balance_after, request.user,
            )
            return redirect(f"{request.path}?date={current.isoformat()}")

        if action == "post":
            context = _leave_ledger_context(current)
            posted = 0
            for row in context["rows"]:
                if row["posted"]:
                    continue
                LeaveLedgerEntry.objects.update_or_create(
                    employee=row["employee"], year=year, month=month,
                    defaults={
                        "full_ot_days": row["full_ot_days"],
                        "el_credited": row["el_credited"],
                        "el_encashed": row["el_encashed"],
                        "el_taken": row["el_taken"],
                        "el_balance_after": row["el_balance_after"],
                        "comp_off_hours_earned": row["comp_off_hours_earned"],
                        "comp_off_balance_after": row["comp_off_balance_after"],
                        "is_manual": False,
                    },
                )
                posted += 1
            messages.success(
                request,
                f"Posted EL/Comp-Off for {posted} Staff employee(s) for {py_calendar.month_name[month]} {year}."
                if posted else "Nothing to post — every Staff employee already has this month posted.",
            )
            logger.info(
                "Leave ledger posted: %s-%s count=%s by user=%s", year, month, posted, request.user,
            )
            return redirect(f"{request.path}?date={current.isoformat()}")

        _error(request, "Unknown action.")
        return redirect(f"{request.path}?date={current.isoformat()}")

    context = _leave_ledger_context(current)
    return render(request, "attendance/leave_ledger.html", context)


def _ot_details_context(date_param: str | None) -> dict:
    """Computes everything the OT page (page and downloads) needs for one
    month — the shift-based OT numbers (see overtime_view) as a Monthly
    Summary grouped by department (one row per employee within each
    department: total OT hours/rate/amount, plus a department header row
    with its subtotal), the read-only Full Monthly View grid, and the
    editable OT View grid (formerly the Attendance dashboard's OT View
    tab — see edit_record_view/bulk_set_shift_view/toggle_month_lock_view,
    which are shared with that dashboard and are agnostic of which page's
    grid triggered them). Shared by ot_details_view and the two download
    views so they can never drift apart."""
    daily_all = _load_daily_data()
    month_keys_all = (
        sorted({(ts.year, ts.month) for ts in daily_all["date"]}) if not daily_all.empty else []
    )
    if not month_keys_all:
        return {"has_data": False}

    default_date = date_cls(*month_keys_all[-1], 1)
    current = _parse_month_date(date_param, default_date)
    year, month = current.year, current.month

    prev_date = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_date = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_nav = {
        "has_data": True,
        "year": year,
        "month": month,
        "month_name": py_calendar.month_name[month],
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
        "current_date": current.isoformat(),
    }

    daily = daily_all[(daily_all["date"].dt.year == year) & (daily_all["date"].dt.month == month)]
    if daily.empty:
        return {"empty": True, "summary_rows": [], **month_nav}

    special_days, downgraded_special_days, skip_special_days = _special_days_and_downgrades()
    daily = metrics.apply_special_days(daily, special_days, downgraded_special_days, skip_special_days)
    emp_rate_map = dict(Employee.objects.values_list("code", "ot_rate_per_hour"))
    full_day_map = _early_closure_hours()
    grid = _build_month_grid(daily, special_days, emp_rate_map, ot_tooltip=True, full_day_map=full_day_map)
    shift_ot_table = grid["shift_ot_table"]
    # Applied to grid["table_rows"] itself (before the deepcopy below) so
    # both tabs show it — a worked Holiday/Paid Holiday/Comp Off should
    # read as "+EL" wherever a Staff employee's month is shown, not just
    # on the editable OT View tab.
    _apply_staff_ot_display(grid["table_rows"], shift_ot_table)
    # Snapshot the full, unrestricted grid for the editable OT View tab
    # before _restrict_to_ot_cells (below) mutates grid["table_rows"] in
    # place for the read-only Full Monthly View tab — same dicts, so
    # without this copy the editable tab would end up with cells blanked
    # out too.
    editable_table_rows = copy.deepcopy(grid["table_rows"])
    # Same Staff Full-OT-day-converts-to-EL exclusion as _build_month_grid
    # (see _ot_payable_table) — applied here too so the Monthly Summary
    # and department breakdown never show a bigger OT total for Staff
    # than what the OT View/Full Monthly View grids above already do.
    staff_codes = (
        set(daily.loc[daily["subcategory"] == "Staff", "emp_code"])
        if "subcategory" in daily.columns else set()
    )
    ot_payable_table = _ot_payable_table(shift_ot_table, staff_codes)
    dept_ot_summary = metrics.department_ot_summary(ot_payable_table)
    is_locked_ot = MonthLock.objects.filter(year=year, month=month, view=MonthLock.VIEW_OT).exists()

    summary_rows = []
    if not ot_payable_table.empty:
        ot_rows = ot_payable_table[ot_payable_table["total_ot_hours"] > 0]
        totals = (
            ot_rows.groupby(["department", "emp_code", "emp_name"])["total_ot_hours"]
            .sum().reset_index()
        )
        emp_rows_by_dept: dict = {}
        for r in totals.itertuples(index=False):
            rate = float(emp_rate_map.get(r.emp_code, 0))
            hours = round(float(r.total_ot_hours), 2)
            emp_rows_by_dept.setdefault(r.department, []).append({
                "is_dept": False,
                "emp_code": r.emp_code,
                "emp_name": r.emp_name,
                "department": r.department,
                "ot_hours": hours,
                # Display-only "Xh Ym" label (see format_hours_as_hm) —
                # "ot_hours" itself stays decimal since the Excel
                # download writes it straight into a worksheet cell.
                "ot_hours_label": metrics.format_hours_as_hm(hours),
                "ot_rate": rate,
                "ot_amount": round(hours * rate, 2),
            })

        for dept in sorted(emp_rows_by_dept, key=metrics.department_sort_key(settings.ATTENDANCE_VISIBLE_DEPARTMENTS)):
            emp_rows = sorted(emp_rows_by_dept[dept], key=lambda row: -row["ot_amount"])
            for i, row in enumerate(emp_rows):
                row["is_top3"] = i < 3
            dept_hours = round(sum(row["ot_hours"] for row in emp_rows), 2)
            dept_amount = round(sum(row["ot_amount"] for row in emp_rows), 2)
            summary_rows.append({
                "is_dept": True,
                "department": dept,
                "headcount": len(emp_rows),
                "ot_hours": dept_hours,
                "ot_hours_label": metrics.format_hours_as_hm(dept_hours),
                "ot_amount": dept_amount,
            })
            summary_rows.extend(emp_rows)

    summary_total_hours = round(sum(r["ot_hours"] for r in summary_rows if not r["is_dept"]), 2)
    # Display-only "Xh Ym" label for the OT View/Full Monthly View tabs'
    # own "Total OT" footer — summary_total_hours itself stays a plain
    # decimal number since the Excel download writes it straight into a
    # worksheet cell (see _ot_details_download_view).
    summary_total_hours_label = metrics.format_hours_as_hm(summary_total_hours)
    summary_total_amount = round(sum(r["ot_amount"] for r in summary_rows if not r["is_dept"]), 2)
    summary_total_el_days = sum(grid["emp_el_days"].values())
    summary_total_permission_hours = metrics.format_hours_as_hm(sum(grid["emp_permission_hours"].values()))
    # Total OT minus Permission Hours, company-wide — same net figure
    # each row's own "paid_ot_hours" shows, summed the same way (from
    # grid["emp_ot_totals"]/emp_permission_hours) so the footer can never
    # drift from what the rows above it add up to.
    summary_total_paid_ot_hours = metrics.format_hours_as_hm(
        sum(grid["emp_ot_totals"].values()) - sum(grid["emp_permission_hours"].values()), allow_negative=True
    )

    # Surfaced on the Monthly Summary tab as a callout — see
    # _is_unconfirmed_special_worked. Computed from editable_table_rows
    # (before _restrict_to_ot_cells mutates the original grid["table_rows"]
    # it was deep-copied from), so this always reflects real punch data
    # regardless of whether the cell ends up visible on Full Monthly View.
    unconfirmed_special_worked = [
        {"label": row["label"].strip(), "date_label": date_cls.fromisoformat(cell["date_iso"]).strftime("%d %b")}
        for row in editable_table_rows if not row["is_dept"]
        for cell in row["day_cells"] if _is_unconfirmed_special_worked(cell)
    ]

    top3_codes = {r["emp_code"] for r in summary_rows if not r["is_dept"] and r.get("is_top3")}
    table_rows = _restrict_to_ot_cells(_ot_only_rows(grid["table_rows"]))
    for row in table_rows:
        if row["is_dept"]:
            continue
        emp_code = next((c["emp_code"] for c in row["day_cells"] if c["emp_code"]), "")
        row["is_top3"] = emp_code in top3_codes

    return {
        **month_nav,
        "empty": False,
        "summary_rows": summary_rows,
        "summary_total_hours": summary_total_hours,
        "summary_total_hours_label": summary_total_hours_label,
        "summary_total_amount": summary_total_amount,
        "summary_total_el_days": summary_total_el_days,
        "summary_total_permission_hours": summary_total_permission_hours,
        "summary_total_paid_ot_hours": summary_total_paid_ot_hours,
        "unconfirmed_special_worked": unconfirmed_special_worked,
        "day_headers": grid["day_headers"],
        "table_rows": table_rows,
        # Not grid["total_cols"] — that's sized for dashboard.html's grid,
        # which renders its own (longer) summary_cols (Work Days/Comp
        # Off/etc.) this report's grid never shows (label + day columns +
        # Total OT/EL Days/Permission Hours/Paid OT Hours/OT Rate/OT
        # Amount only).
        "total_cols": 1 + len(grid["day_headers"]) + 6,
        "heat_colors": metrics.HEAT_COLORS,
        # For the editable OT View tab (moved here from the Attendance
        # dashboard) — the full roster (not just employees with OT this
        # month), so a fresh OT shift can still be assigned to anyone.
        "editable_table_rows": editable_table_rows,
        "dept_ot_summary": dept_ot_summary,
        "is_locked_ot": is_locked_ot,
        "day_types": SpecialDay.TYPE_CHOICES,
        "status_choices": AttendanceRecord.STATUS_CHOICES,
    }


@login_required
def ot_details_view(request):
    """OT page — see _ot_details_context for the actual computation.
    Renders as three tabs: OT View (the editable day x employee grid,
    moved here from the Attendance dashboard), Monthly Summary (grouped
    by department), and Full Monthly View (the read-only day x employee
    grid), scoped to one calendar month with prev/next nav."""
    context = _ot_details_context(request.GET.get("date"))
    return render(request, "attendance/ot_details.html", context)


def _apply_grid_borders(ws, min_row: int, max_row: int, max_col: int) -> None:
    """Thin border on every side of every cell in the given range — used
    by the OT Details .xlsx downloads so the header/data/total rows read
    as a proper ruled table instead of borderless openpyxl default cells."""
    from openpyxl.styles import Border, Side

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border


def _write_ot_summary_sheet(ws, context) -> None:
    """Fills in the OT Monthly Summary sheet (grouped by department, one
    row per employee: OT hours, OT rate, OT amount, plus a department
    subtotal row) — same numbers and colors as the report page's Monthly
    Summary tab: department rows get the same peach fill as the page's
    tr.dept-row, and the top 3 OT earners in each department (already
    sorted first, see _ot_details_context) get the page's light blue fill
    and bold text."""
    from openpyxl.styles import Font, PatternFill
    import openpyxl

    dept_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    top3_fill = PatternFill(start_color="EAF4FC", end_color="EAF4FC", fill_type="solid")

    month_label = f"{context['month_name']} {context['year']}"
    ws.append([f"OT Monthly Summary — {month_label}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    headers = ["Code", "Employee", "Department", "OT Hours", "OT Rate/hr", "OT Amount"]
    ws.append(headers)
    header_row_num = ws.max_row
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in context["summary_rows"]:
        if row["is_dept"]:
            # Dept OT Hours/Amount used to be their own trailing columns
            # (blank on every employee row) — folded into the department
            # row's own label instead, next to the headcount.
            label = f"{row['department']} ({row['headcount']}) — {row['ot_hours']} hrs, {row['ot_amount']}"
            ws.append([label])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = dept_fill
        else:
            ws.append([
                row["emp_code"], row["emp_name"], row["department"],
                row["ot_hours"], row["ot_rate"], row["ot_amount"],
            ])
            if row.get("is_top3"):
                for cell in ws[ws.max_row]:
                    cell.fill = top3_fill
                    cell.font = Font(bold=True)

    ws.append(["", "", "Total", context["summary_total_hours"], "", context["summary_total_amount"]])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF9B0", end_color="FFF9B0", fill_type="solid")

    _apply_grid_borders(ws, header_row_num, ws.max_row, len(headers))

    widths = [10, 26, 22, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def _write_ot_grid_sheet(ws, context) -> None:
    """Fills in the OT Full Monthly View sheet (day x employee grid,
    grouped by department, OT-relevant cells only — see
    _restrict_to_ot_cells). Every hours/rate/amount column is written as a
    real number (not the HTML grid's display-formatted strings), header
    cells get a grey fill, department rows get the same peach fill as the
    page's tr.dept, day cells get the page's Holiday/Paid Holiday/Comp Off
    colors or its confirmed-OT-shift yellow, the Total OT column gets a
    light grey fill on every row, the top 3 OT earners per department
    (row["is_top3"], set in _ot_details_context by matching summary_rows'
    ranking to each row's employee code) get the same light blue fill +
    bold as the Summary sheet — overriding every other fill on that row,
    same simplification the Total OT column's grey already made — and a
    bold Total row sums Total OT/OT Amount across all employees; grid
    borders cover the whole table including that row."""
    from openpyxl.styles import Font, PatternFill
    import openpyxl

    dept_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    top3_fill = PatternFill(start_color="EAF4FC", end_color="EAF4FC", fill_type="solid")
    total_ot_col_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    special_day_fills = {
        "H": PatternFill(start_color="D6D6D6", end_color="D6D6D6", fill_type="solid"),
        "PH": PatternFill(start_color="4FB3A9", end_color="4FB3A9", fill_type="solid"),
        "CO": PatternFill(start_color="F5A85C", end_color="F5A85C", fill_type="solid"),
    }
    confirmed_ot_fill = PatternFill(start_color="FFF9B0", end_color="FFF9B0", fill_type="solid")

    month_label = f"{context['month_name']} {context['year']}"
    day_headers = context["day_headers"]
    ws.append([f"OT Full Monthly View — {month_label}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    # day_headers' "label" is a display string ("1", "2", ... the
    # day-of-month) — write it as a real number, same fix as the day cells
    # below, so the header row isn't text either.
    day_number_headers = [
        int(d["label"]) if d["label"].isdigit() else d["label"] for d in day_headers
    ]
    headers = ["Employee"] + day_number_headers + ["Total OT", "OT Rate/hr", "OT Amount"]
    ws.append(headers)
    header_row_num = ws.max_row
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    total_ot_col = headers.index("Total OT") + 1

    total_ot_sum = Decimal(0)
    total_amount_sum = Decimal(0)
    for row in context["table_rows"]:
        if row["is_dept"]:
            ws.append([row["label"].lstrip("▸ ").strip()])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = dept_fill
        else:
            # cell["value"] is the day's full worked hours; the on-screen
            # grid immediately swaps that to cell["shift_ot_hours"] (the
            # real per-day OT credit) via JS for every non-blank cell (see
            # the "Full Monthly View grid" script below) — do the same
            # substitution here so the sheet matches what's actually shown
            # on screen instead of the full work-hours fallback.
            day_values = [
                float(cell["shift_ot_hours"]) if cell["shift_ot_hours"] != "" else None
                for cell in row["day_cells"]
            ]
            ws.append(
                [row["label"].strip()] + day_values
                + [row["total_ot"] or None, row["ot_rate"] or None, row["total_ot_amount"] or None]
            )
            if row["total_ot"]:
                total_ot_sum += Decimal(str(row["total_ot"]))
            if row["total_ot_amount"]:
                total_amount_sum += Decimal(str(row["total_ot_amount"]))

            data_row_num = ws.max_row
            for i, cell in enumerate(row["day_cells"]):
                day_fill = special_day_fills.get(cell["special"]) or (confirmed_ot_fill if cell["shift_flag"] else None)
                if day_fill:
                    ws.cell(row=data_row_num, column=2 + i).fill = day_fill
            if row.get("is_top3"):
                for cell in ws[data_row_num]:
                    cell.fill = top3_fill
                    cell.font = Font(bold=True)
            else:
                ws.cell(row=data_row_num, column=total_ot_col).fill = total_ot_col_fill

    ws.append(
        ["Total"] + [None] * len(day_headers) + [float(total_ot_sum), None, float(total_amount_sum)]
    )
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF9B0", end_color="FFF9B0", fill_type="solid")

    _apply_grid_borders(ws, header_row_num, ws.max_row, len(headers))

    ws.column_dimensions["A"].width = 26
    for i in range(2, 2 + len(day_headers)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5
    for i in range(2 + len(day_headers), 5 + len(day_headers)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12


@login_required
def ot_details_download_view(request):
    """Downloads both OT Details views for one month as a single .xlsx with
    two sheets — "OT Summary" (Monthly Summary tab) and "OT Full Monthly
    View" (day x employee grid) — via the same _ot_details_context used by
    the report page, so the two sheets and the on-screen tabs can never
    drift apart."""
    import openpyxl

    context = _ot_details_context(request.GET.get("date"))
    if not context.get("has_data") or context.get("empty"):
        raise Http404("No OT data for that month.")

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "OT Summary"
    _write_ot_summary_sheet(summary_ws, context)

    grid_ws = wb.create_sheet("OT Full Monthly View")
    _write_ot_grid_sheet(grid_ws, context)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"ot-details-{context['year']}-{context['month']:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
