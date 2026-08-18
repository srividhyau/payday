"""
Populates Employee bank-transfer details (account name/no, bank, IFSC,
branch) from the monthly salary workbook's "Bank Details" sheet (e.g.
"July Salary.xlsm"). That workbook is a different source than the eSSL
DailyAttendance export (see importer.py) — it's HR/payroll's own sheet.

"Bank Details" stacks several sections top to bottom (Operators, an
unlabeled block, Ironing & Bartrack, Helpers, Company Workers,
Fixed_Payment), each with its own repeated header row (Emp Name, Account
Name, Bank Name, Account No, IFSC Code, Branch) and no employee code column
at all — so every match here is by exact name, not by the more reliable
Employee.code used elsewhere in the attendance system.

Only updates *existing* Employee rows. Never creates new employees —
department/category for a genuinely new hire isn't reliably inferable from
this sheet, so unmatched rows are reported instead, for HR to reconcile by
hand (usually a name spelling mismatch against the attendance system).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

from .models import Department, Employee

SHEET_NAME = "Bank Details"

# Department sheets: same workbook, one tab per pay group. Unlike Bank
# Details, three of these (Op/I&B/Staff) have no per-employee Dept column at
# all — only Helpers and Company Workers do — so those three get a fixed
# label per sheet instead of a per-row value.
DEPARTMENT_SHEETS: dict[str, str | None] = {
    "Op": "Operators",
    "I&B": "Ironing & Bartrack",
    "Staff": "Staff",
    "Helpers": None,  # per-row DEPT column
    "Company Workers": None,  # per-row Dept column
}

_HEADER_ALIASES = {
    "emp name": "emp_name",
    "account name": "account_name",
    "bank name": "bank_name",
    "account no": "account_no",
    "ifsc code": "ifsc_code",
    "branch": "branch",
}

_BANK_FIELDS = ("account_name", "bank_name", "account_no", "ifsc_code", "branch")

_DEPT_HEADER_ALIASES = {
    "emp name": "emp_name",
    "name": "name",
    "emp no": "emp_no",
    "dept": "dept",
}


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


@dataclass
class ImportReport:
    updated: list[str] = field(default_factory=list)
    unmatched: list[str] = field(default_factory=list)
    conflicts: list[str] = field(default_factory=list)
    created: list[str] = field(default_factory=list)


def _find_header_rows(rows: list[tuple]) -> list[tuple[int, dict[str, int]]]:
    """Every row containing an 'Emp Name' cell starts a new section — the
    sheet repeats this header once per department block."""
    headers = []
    for row_idx, row in enumerate(rows):
        col_map = {}
        for col_idx, cell in enumerate(row):
            field_name = _HEADER_ALIASES.get(_normalize_header(cell))
            if field_name and field_name not in col_map:
                col_map[field_name] = col_idx
        if "emp_name" in col_map:
            headers.append((row_idx, col_map))
    return headers


def _extract_bank_fields(row: tuple, col_map: dict[str, int]) -> dict[str, str]:
    values = {}
    for field_name in _BANK_FIELDS:
        col = col_map.get(field_name)
        if col is not None and col < len(row):
            values[field_name] = _clean(row[col])
    return values


def _generate_placeholder_code(name: str) -> str:
    """Bank Details has no employee code column, so a newly-created employee
    needs a stand-in. This is NOT a real eSSL code — if this person later
    appears in a genuine attendance upload under their true code, that will
    create a second Employee row, and someone will need to reconcile the two
    by hand (e.g. move this record's bank details onto the correctly-coded
    one and delete this one)."""
    base = re.sub(r"[^A-Z0-9]+", "", name.upper())[:24]
    code = base or "EMP"
    suffix = 2
    while Employee.objects.filter(code=code).exists():
        code = f"{base}-{suffix}"
        suffix += 1
    return code


def import_bank_details(path, dry_run: bool = False, create_missing: bool = False) -> ImportReport:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No '{SHEET_NAME}' sheet found in {path}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    headers = _find_header_rows(rows)
    report = ImportReport()
    # Same name can appear in more than one section with *different* bank
    # data (two different people sharing a first name, or a stray
    # duplicate) — since matching here is name-only, silently letting the
    # second row overwrite the first risks assigning the wrong bank account
    # for a real money transfer. Track what's already been applied per
    # employee this run and flag a conflict instead of overwriting.
    applied: dict[str, tuple[str, dict[str, str]]] = {}
    # In dry-run mode nothing is actually persisted, so a name created
    # earlier in this same run wouldn't be found by a fresh DB query on a
    # later occurrence — cache it locally so conflict detection between two
    # rows sharing a *new* name works identically in dry-run and real runs.
    created_this_run: dict[str, Employee] = {}

    for block_num, (header_idx, col_map) in enumerate(headers):
        name_col = col_map["emp_name"]
        end_idx = headers[block_num + 1][0] if block_num + 1 < len(headers) else len(rows)

        for row in rows[header_idx + 1:end_idx]:
            name_val = row[name_col] if name_col < len(row) else None
            if not isinstance(name_val, str) or not name_val.strip():
                continue

            name = name_val.strip()
            bank_fields = _extract_bank_fields(row, col_map)
            # Section-title rows ("Ironing & Bartrack Bank", a stray repeated
            # header, etc.) land in this same column range with no bank data
            # at all — skip them without reporting, they aren't employees.
            if not any(bank_fields.values()):
                continue

            emp = Employee.objects.filter(name__iexact=name).first() or created_this_run.get(name.lower())
            created_now = False
            if emp is None:
                if not create_missing:
                    report.unmatched.append(name)
                    continue
                code = _generate_placeholder_code(name)
                if dry_run:
                    # Can't call .create() in dry-run without persisting, so
                    # build an unsaved instance purely to report against —
                    # the code shown is illustrative, the real run may pick
                    # a different suffix if other rows create employees first.
                    emp = Employee(code=code, name=name)
                else:
                    emp = Employee.objects.create(code=code, name=name)
                created_now = True
                created_this_run[name.lower()] = emp
                report.created.append(f"{code} {name}")

            prior = applied.get(emp.code)
            if prior is not None:
                prior_name, prior_fields = prior
                if any(
                    bank_fields.get(k) and prior_fields.get(k) and bank_fields[k] != prior_fields[k]
                    for k in _BANK_FIELDS
                ):
                    report.conflicts.append(
                        f"{emp.code} {emp.name}: '{prior_name}' gave "
                        f"{prior_fields.get('account_no')!r}/{prior_fields.get('bank_name')!r}, "
                        f"'{name}' gives {bank_fields.get('account_no')!r}/{bank_fields.get('bank_name')!r} "
                        "— skipped, resolve manually"
                    )
                continue

            if created_now:
                for field_name, value in bank_fields.items():
                    if value:
                        setattr(emp, field_name, value)
                if not dry_run:
                    emp.save()
            else:
                changed = []
                for field_name, value in bank_fields.items():
                    if value and getattr(emp, field_name) != value:
                        setattr(emp, field_name, value)
                        changed.append(field_name)
                if changed:
                    if not dry_run:
                        emp.save(update_fields=changed)
                    report.updated.append(f"{emp.code} {emp.name} (via name '{name}': {', '.join(changed)})")

            applied[emp.code] = (name, bank_fields)

    return report


@dataclass
class DeptImportReport:
    updated: list[str] = field(default_factory=list)
    unmatched: list[str] = field(default_factory=list)


def _find_dept_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """First row containing an 'EMP NAME' cell — each department sheet
    combines its salary table and (where present) bank columns into one
    header row, same as the Bank Details sheet's per-section headers."""
    for row_idx, row in enumerate(rows):
        col_map = {}
        for col_idx, cell in enumerate(row):
            field_name = _DEPT_HEADER_ALIASES.get(_normalize_header(cell))
            if field_name and field_name not in col_map:
                col_map[field_name] = col_idx
        if "emp_name" in col_map:
            return row_idx, col_map
    return None


def _resolve_department(name: str) -> Department:
    """Matches against an existing Department first — case-insensitively,
    and tolerating the "X" vs "X DEPARTMENT" naming mismatch between this
    workbook's Dept column (e.g. "CUTTING") and the eSSL attendance import's
    Department names (e.g. "CUTTING DEPARTMENT") — before creating a new
    one, so the two sources don't end up with duplicate near-identical rows."""
    name = name.strip()
    dept = Department.objects.filter(name__iexact=name).first()
    if dept:
        return dept
    dept = Department.objects.filter(name__iexact=f"{name} DEPARTMENT").first()
    if dept:
        return dept
    for candidate in Department.objects.all():
        if candidate.name.upper().replace(" DEPARTMENT", "").strip() == name.upper():
            return candidate
    return Department.objects.create(name=name)


def import_departments(path, dry_run: bool = False) -> DeptImportReport:
    """Fills in Employee.department from the salary workbook's department
    sheets — but only where it's currently unset. This field's primary
    source of truth is the eSSL attendance import (importer.py), which
    generally has more precise/consistent names; this only backfills gaps,
    it never overwrites an existing value, so it can't regress data that's
    already correct."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    report = DeptImportReport()

    for sheet_name, fixed_label in DEPARTMENT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

        header = _find_dept_header_row(rows)
        if header is None:
            continue
        header_idx, col_map = header
        name_col = col_map.get("emp_name", col_map.get("name"))
        emp_no_col = col_map.get("emp_no")
        dept_col = col_map.get("dept")

        for row in rows[header_idx + 1:]:
            name_val = row[name_col] if name_col < len(row) else None
            if not isinstance(name_val, str) or not name_val.strip():
                break
            name = name_val.strip()

            dept_label = fixed_label
            if dept_col is not None and dept_col < len(row):
                cell_val = _clean(row[dept_col])
                if cell_val:
                    dept_label = cell_val
            if not dept_label:
                continue

            emp = None
            if emp_no_col is not None and emp_no_col < len(row):
                code = _clean(row[emp_no_col])
                if code:
                    emp = Employee.objects.filter(code=code).first()
            if emp is None:
                emp = Employee.objects.filter(name__iexact=name).first()
            if emp is None:
                report.unmatched.append(f"{sheet_name}: {name}")
                continue

            if emp.department is not None:
                continue

            dept = _resolve_department(dept_label)
            if not dry_run:
                emp.department = dept
                emp.save(update_fields=["department"])
            report.updated.append(f"{emp.code} {emp.name} -> {dept.name} (from {sheet_name})")

    return report
